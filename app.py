import os
import re
import json

def safe_json(obj):
    return json.dumps(obj).replace("</", "<\\/")

import secrets
import hmac
import time
from datetime import datetime, timedelta
from functools import wraps

from flask import (
    Flask, flash, redirect, render_template, request, session,
    url_for, abort, send_from_directory, Response
)

from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash

from sqlalchemy import inspect as sa_inspect

from authlib.integrations.flask_client import OAuth
from werkzeug.middleware.proxy_fix import ProxyFix

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

BASE_DIR = os.path.abspath(os.path.dirname(__file__))

app = Flask(
    __name__,
    template_folder=os.path.join(BASE_DIR, "templates"),
    static_folder=os.path.join(BASE_DIR, "static")
)

app.wsgi_app = ProxyFix(app.wsgi_app, x_proto=1, x_host=1)

IS_PRODUCTION = bool(os.environ.get("VERCEL")) or bool(os.environ.get("DATABASE_URL"))

_secret_key_env = os.environ.get("SECRET_KEY")

if not _secret_key_env and IS_PRODUCTION:
    print(
        "[HOSTEL APP] WARNING: SECRET_KEY is not set in a production environment. "
        "Sessions will be invalidated on every restart/cold-start and may not be "
        "shared correctly across serverless instances. Set SECRET_KEY in your "
        "environment variables."
    )

app.secret_key = _secret_key_env or secrets.token_hex(32)

DATABASE_URL = os.environ.get("DATABASE_URL")
BASE_DIR = os.path.abspath(os.path.dirname(__file__))

app = Flask(
    __name__,
    template_folder=os.path.join(BASE_DIR, "templates"),
    static_folder=os.path.join(BASE_DIR, "static")
)

app.wsgi_app = ProxyFix(app.wsgi_app, x_proto=1, x_host=1)

IS_PRODUCTION = bool(os.environ.get("VERCEL")) or bool(os.environ.get("DATABASE_URL"))

_secret_key_env = os.environ.get("SECRET_KEY")

if not _secret_key_env and IS_PRODUCTION:
    print(
        "[HOSTEL APP] WARNING: SECRET_KEY is not set in a production environment. "
        "Sessions will be invalidated on every restart/cold-start and may not be "
        "shared correctly across serverless instances. Set SECRET_KEY in your "
        "environment variables."
    )

app.secret_key = _secret_key_env or secrets.token_hex(32)

DATABASE_URL = os.environ.get("DATABASE_URL")

if DATABASE_URL:
    if DATABASE_URL.startswith("postgres://"):
        DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)

    app.config["SQLALCHEMY_DATABASE_URI"] = DATABASE_URL

    app.config["SQLALCHEMY_ENGINE_OPTIONS"] = {
        "pool_pre_ping": True,
        "pool_recycle": 300,
    }

elif os.environ.get("VERCEL"):
    os.makedirs("/tmp", exist_ok=True)

    app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:////tmp/app.sqlite3"

    print(
        "[HOSTEL APP] WARNING: DATABASE_URL is not set. Using a temporary SQLite "
        "database in /tmp so the app can still run, but ALL DATA WILL BE LOST on "
        "every cold start/redeploy. Set DATABASE_URL to a real Postgres connection "
        "string in your Vercel project's environment variables for real persistence."
    )

else:
    app.config["SQLALCHEMY_DATABASE_URI"] = (
        "sqlite:///" + os.path.join(BASE_DIR, "app.sqlite3")
    )

    print("WARNING: Using SQLite database - data will not persist on server restart!")

app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["MAX_CONTENT_LENGTH"] = 10 * 1024 * 1024
app.config["SESSION_COOKIE_SECURE"] = IS_PRODUCTION
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(hours=24)
db = SQLAlchemy(app)
 
oauth = OAuth(app)
google_oauth = oauth.register(
    name="google",
    client_id=os.environ.get("GOOGLE_CLIENT_ID"),
    client_secret=os.environ.get("GOOGLE_CLIENT_SECRET"),
    server_metadata_url="https://accounts.google.com/.well-known/openid-configuration",
    client_kwargs={
        "scope": "openid email profile"
    },
)
if os.environ.get("VERCEL"):
    UPLOAD_DIR = os.path.join("/tmp", "uploads")
else:
    UPLOAD_DIR = os.path.join(BASE_DIR, "static", "uploads")
DB_READY = False
 
ADMIN_EMAIL = "admin@sece.ac.in"
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "Admin@1234")
if ADMIN_PASSWORD == "Admin@1234" and IS_PRODUCTION:
    print("[HOSTEL APP] WARNING: ADMIN_PASSWORD is not set — the admin account is using "
          "the default password. Set ADMIN_PASSWORD in your environment variables.")
 
VALID_CATEGORIES = {"Electrical", "Plumbing", "Wi-Fi", "Cleaning", "Furniture", "Hostel food", "Others"}
VALID_PRIORITIES = {"Low", "Moderate", "High"}
 
try:
    from google import genai
    gemini_api_key = os.environ.get("GEMINI_API_KEY")
    if gemini_api_key:
        genai_client = genai.Client(api_key=gemini_api_key)
    else:
        print("[HOSTEL APP] Warning: GEMINI_API_KEY environment variable not set.")
        genai_client = None
except Exception as e:
    print(f"[HOSTEL APP] Warning: Could not initialize Gemini client: {e}")
    genai_client = None
 
import threading
 
def classify_priority_async(app_instance, complaint_id, description):
    def run():
        with app_instance.app_context():
            try:
                if not genai_client:
                    return
                prompt = (
                    "Determine the priority level of the following hostel complaint description. "
                    "Respond with exactly one word: 'High', 'Moderate', or 'Low'. "
                    "Do not include any other text, explanation, or punctuation.\n\n"
                    f"Complaint Description: {description}"
                )
                response = genai_client.models.generate_content(
                    model="gemini-2.0-flash",
                    contents=prompt
                )
                val = response.text.strip().capitalize()
                if val in {"High", "Moderate", "Low"}:
                    complaint = db.session.get(Complaint, complaint_id)
                    if complaint:
                        complaint.priority = val
                        db.session.commit()
                        print(f"[HOSTEL APP] Updated complaint #{complaint_id} priority to {val} via Gemini async.")
            except Exception as ex:
                print(f"[HOSTEL APP] Async Gemini classification failed for complaint #{complaint_id}: {ex}")
 
    threading.Thread(target=run, daemon=True).start()
VALID_STATUSES = {"Pending", "In Progress", "Resolved"}
VALID_HOSTEL_TYPES = {"boys", "girls"}
VALID_BLOCKS = {"A", "B", "C", "D", "E", "F"}
 
 
def sanitize_string(text, max_length=255):
    if not text:
        return ""
    text = str(text).strip()
    text = re.sub(r'[<>\"\'%;()&+]', '', text)
    return text[:max_length]
 
 
def sanitize_description(text, max_length=500):
    if not text:
        return ""
    text = str(text).strip()
    text = re.sub(r'<[^>]+>', '', text)
    return text[:max_length]
 
 
def verify_and_maybe_upgrade_password(profile, raw_password):
    """
    Checks a submitted password against a StudentProfile's stored password.

    Existing rows created before password hashing was introduced still hold
    plaintext values, so this transparently upgrades them to a secure hash
    the moment the student successfully logs in — no forced reset, no data
    loss, no broken accounts. New/updated rows are always stored hashed via
    hash_password() below.
    """
    stored = profile.password or ""
    if not stored:
        return False
    if stored.startswith(("pbkdf2:", "scrypt:")):
        return check_password_hash(stored, raw_password)
    # Legacy plaintext row: constant-time compare, then upgrade in place.
    if hmac.compare_digest(stored, raw_password):
        profile.password = generate_password_hash(raw_password)
        db.session.commit()
        return True
    return False


def hash_password(raw_password):
    return generate_password_hash(raw_password)


# --- Login rate limiting -----------------------------------------------
# Simple in-process sliding-window limiter. Keyed by IP+email so one
# student mistyping their password can't lock out the whole hostel, and a
# shared campus IP can't be used to lock out a single account either.
_login_attempts = {}
LOGIN_MAX_ATTEMPTS = 6
LOGIN_WINDOW_SECONDS = 300  # 5 minutes


def _login_attempt_key(email):
    return f"{request.remote_addr}:{email}"


def _too_many_login_attempts(key):
    now = time.time()
    attempts = [t for t in _login_attempts.get(key, []) if now - t < LOGIN_WINDOW_SECONDS]
    _login_attempts[key] = attempts
    return len(attempts) >= LOGIN_MAX_ATTEMPTS


def _record_failed_login(key):
    _login_attempts.setdefault(key, []).append(time.time())


def _clear_login_attempts(key):
    _login_attempts.pop(key, None)


COMPLAINTS_PAGE_SIZE = 15


def apply_complaint_filters(query, args):
    """Server-side filtering for the complaint table — real DB-level
    filtering (not a client-side scan of already-loaded rows)."""
    search = (args.get("q") or "").strip()
    status = (args.get("status") or "").strip()
    category = (args.get("category") or "").strip()
    priority = (args.get("priority") or "").strip()
    hostel = (args.get("hostel") or "").strip().lower()
    block = (args.get("block") or "").strip().upper()
    date_from = (args.get("date_from") or "").strip()
    date_to = (args.get("date_to") or "").strip()
    student_name = (args.get("student_name") or "").strip()
    room = (args.get("room") or "").strip()
    warden_id = (args.get("warden") or "").strip()

    if search:
        like = f"%{search[:120]}%"
        query = query.filter(db.or_(
            Complaint.student_name.ilike(like),
            Complaint.room_number.ilike(like),
            Complaint.category.ilike(like),
            Complaint.description.ilike(like),
            Complaint.block.ilike(like),
        ))
    if student_name:
        query = query.filter(Complaint.student_name.ilike(f"%{student_name[:120]}%"))
    if room:
        query = query.filter(Complaint.room_number.ilike(f"%{room[:50]}%"))
    if status and status.lower() != "all":
        query = query.filter(Complaint.status == status)
    if category:
        query = query.filter(Complaint.category.ilike(category))
    if priority:
        query = query.filter(Complaint.priority.ilike(priority))
    if hostel and hostel != "all":
        query = query.filter(db.func.lower(Complaint.hostel_type) == hostel)
    if block and block != "ALL":
        query = query.filter(db.func.upper(Complaint.block) == block)
    if warden_id and warden_id.isdigit():
        w = Warden.query.get(int(warden_id))
        if w:
            query = query.filter(
                db.func.lower(Complaint.hostel_type) == w.hostel_type.lower(),
                db.func.upper(Complaint.block) == w.block.upper(),
            )
    if date_from:
        try:
            query = query.filter(Complaint.created_at >= datetime.strptime(date_from, "%Y-%m-%d"))
        except ValueError:
            pass
    if date_to:
        try:
            query = query.filter(Complaint.created_at < datetime.strptime(date_to, "%Y-%m-%d") + timedelta(days=1))
        except ValueError:
            pass
    return query


def _priority_rank_col():
    return db.case(
        (Complaint.priority.ilike("high"), 3),
        (Complaint.priority.ilike("moderate"), 2),
        (Complaint.priority.ilike("low"), 1),
        else_=0,
    )


def _status_rank_col():
    return db.case(
        (Complaint.status == "Pending", 1),
        (Complaint.status == "In Progress", 2),
        (Complaint.status == "Resolved", 3),
        else_=4,
    )


def apply_complaint_sort(query, sort_by):
    """Accepts either a legacy single-word value from the old <select>
    ('latest', 'oldest', 'priority', 'status') or a comma-separated list of
    '<field>_<asc|desc>' tokens driven by clickable, shift-click-for-
    secondary table headers (e.g. 'priority_desc,date_desc'). Always falls
    back to newest-first, and always adds created_at as a final tiebreaker
    so multi-column sorts are stable."""
    sort_by = (sort_by or "latest").strip()

    sort_field_columns = {
        "date": Complaint.created_at,
        "name": Complaint.student_name,
        "room": Complaint.room_number,
        "category": Complaint.category,
        "block": Complaint.block,
    }

    legacy = {
        "latest": [("date", "desc")],
        "oldest": [("date", "asc")],
        "priority": [("priority", "desc")],
        "status": [("status", "asc")],
    }
    if sort_by in legacy:
        tokens = legacy[sort_by]
    else:
        tokens = []
        for part in sort_by.split(","):
            part = part.strip()
            if "_" not in part:
                continue
            field, _, direction = part.rpartition("_")
            if field and direction in ("asc", "desc"):
                tokens.append((field, direction))
        if not tokens:
            tokens = [("date", "desc")]

    order_clauses = []
    seen_fields = set()
    for field, direction in tokens[:3]:  # cap at 3 sort keys, sane limit
        if field in seen_fields:
            continue
        seen_fields.add(field)
        if field == "priority":
            col = _priority_rank_col()
        elif field == "status":
            col = _status_rank_col()
        elif field in sort_field_columns:
            col = sort_field_columns[field]
        else:
            continue
        order_clauses.append(col.asc() if direction == "asc" else col.desc())

    order_clauses.append(Complaint.created_at.desc())
    return query.order_by(*order_clauses)


def _csv_safe(value):
    """
    Neutralizes CSV/formula injection: a description or note starting with
    =, +, -, or @ would be interpreted as a formula by Excel/Sheets when
    the exported CSV is opened, potentially executing attacker-controlled
    logic on whoever opens the report. Prefixing with a leading apostrophe
    forces spreadsheet apps to treat it as plain text instead.
    """
    text = "" if value is None else str(value)
    if text and text[0] in ("=", "+", "-", "@"):
        return "'" + text
    return text


def determine_priority(description):
    # Local keyword-based priority determination as fallback
    desc_lower = (description or "").lower()
    
    # Critical/urgent issues: safety hazards, major water leaks, complete power/internet failure
    high_keywords = [
        "emergency", "shock", "spark", "short circuit", "fire", "flood", 
        "burst", "no water", "power cut", "power outage", "current", "wire",
        "broken pipe", "blockage", "toilet", "stink", "food poisoning", "poisoning",
        "snake", "urgent", "danger", "hazard", "leakage", "overflow", "injured",
        "accident", "broken glass", "shattered", "lockout", "locked out", "theft", "stolen"
    ]
    
    # Minor, slow, or non-blocking issues
    low_keywords = [
        "slow", "speed", "signal", "dusty", "dirty", "clean", "furniture", 
        "mirror", "paint", "bulb", "fan slow", "wifi slow", "dust"
    ]
    
    fallback_priority = "Moderate"
    for kw in high_keywords:
        if kw in desc_lower:
            fallback_priority = "High"
            break
    if fallback_priority != "High":
        for kw in low_keywords:
            if kw in desc_lower:
                fallback_priority = "Low"
                break
            
    return fallback_priority
 
 
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "user" not in session:
            flash("Please login first.", "error")
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated_function
 
 
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get("role") == "admin":
            abort(403)
        return f(*args, **kwargs)
    return decorated_function
 
 
class Complaint(db.Model):
    __tablename__ = "complaints"
    id = db.Column(db.Integer, primary_key=True)
    student_email = db.Column(db.String(120), nullable=False, index=True)
    student_name = db.Column(db.String(120), nullable=False)
    room_number = db.Column(db.String(50), nullable=False)
    floor = db.Column(db.String(10), nullable=True)
    category = db.Column(db.String(50), nullable=False, index=True)
    priority = db.Column(db.String(10), nullable=False, index=True)
    description = db.Column(db.Text, nullable=False)
    hostel_type = db.Column(db.String(10), nullable=True, index=True)
    block = db.Column(db.String(5), nullable=True, index=True)
    image_filename = db.Column(db.String(255), nullable=True)
    status = db.Column(db.String(20), nullable=False, default="Pending", index=True)
    admin_note = db.Column(db.Text, nullable=True)
    created_at = db.Column(db.DateTime, nullable=False, default=datetime.utcnow, index=True)
    updated_at = db.Column(db.DateTime, nullable=False, default=datetime.utcnow, onupdate=datetime.utcnow, index=True)
 
    def to_dict(self):
        return {
            "id": self.id,
            "student_email": self.student_email,
            "student_name": self.student_name,
            "room_number": self.room_number,
            "floor": self.floor,
            "category": self.category,
            "priority": self.priority,
            "description": self.description,
            "hostel_type": self.hostel_type,
            "block": self.block,
            "image_filename": self.image_filename,
            "status": self.status,
            "admin_note": self.admin_note,
            "created_at": self.created_at.isoformat() if self.created_at else None,
            "updated_at": self.updated_at.isoformat() if self.updated_at else None,
        }
 
 
class IssueCounter(db.Model):
    __tablename__ = "issue_counter"
    id = db.Column(db.Integer, primary_key=True)
    total = db.Column(db.Integer, nullable=False, default=0)
 
    @classmethod
    def get(cls):
        row = cls.query.first()
        if not row:
            row = cls(total=0)
            db.session.add(row)
            db.session.commit()
        return row
 
 
class StudentProfile(db.Model):
    __tablename__ = "student_profiles"
    email = db.Column(db.String(120), primary_key=True)
    name = db.Column(db.String(120), nullable=False)
    roll_number = db.Column(db.String(50), nullable=True)
    dept = db.Column(db.String(100), nullable=True)
    year = db.Column(db.String(10), nullable=True)
    phone = db.Column(db.String(20), nullable=True)
    password = db.Column(db.String(255), nullable=True)
    hostel_type = db.Column(db.String(10), nullable=True)
    block = db.Column(db.String(10), nullable=True)
    floor = db.Column(db.String(10), nullable=True)
    room_no = db.Column(db.String(20), nullable=True)
 
 
class Warden(db.Model):
    __tablename__ = "wardens"
    id = db.Column(db.Integer, primary_key=True)
    hostel_type = db.Column(db.String(10), nullable=False) # 'boys' or 'girls'
    block = db.Column(db.String(5), nullable=False)        # 'A', 'B', 'C', 'D', 'E', 'F'
    name = db.Column(db.String(120), nullable=False)
    contact = db.Column(db.String(50), nullable=False)
 
    def to_dict(self):
        return {
            "id": self.id,
            "hostel_type": self.hostel_type,
            "block": self.block,
            "name": self.name,
            "contact": self.contact
        }
 
 
class Notice(db.Model):
    __tablename__ = "notices"
    id = db.Column(db.Integer, primary_key=True)
    hostel_type = db.Column(db.String(10), nullable=False) # 'boys', 'girls', or 'all'
    block = db.Column(db.String(10), nullable=False)       # 'A', 'B', 'C', 'D', 'E', 'F', or 'all'
    title = db.Column(db.String(120), nullable=False)
    content = db.Column(db.Text, nullable=False)
    created_at = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    resolved_by = db.Column(db.DateTime, nullable=True)
 
    def to_dict(self):
        return {
            "id": self.id,
            "hostel_type": self.hostel_type,
            "block": self.block,
            "title": self.title,
            "content": self.content,
            "created_at": self.created_at.strftime('%Y-%m-%d %H:%M:%S') if self.created_at else "",
            "resolved_by": self.resolved_by.strftime('%Y-%m-%d %H:%M:%S') if self.resolved_by else ""
        }


class Notification(db.Model):
    __tablename__ = "notifications"
    id = db.Column(db.Integer, primary_key=True)
    recipient_email = db.Column(db.String(120), nullable=False, index=True)
    recipient_role = db.Column(db.String(20), nullable=False, default="student")  # 'student' or 'admin'
    type = db.Column(db.String(40), nullable=False, default="general")
    title = db.Column(db.String(150), nullable=False)
    message = db.Column(db.String(500), nullable=False)
    link = db.Column(db.String(255), nullable=True)
    is_read = db.Column(db.Boolean, nullable=False, default=False, index=True)
    created_at = db.Column(db.DateTime, nullable=False, default=datetime.utcnow, index=True)

    def to_dict(self):
        return {
            "id": self.id,
            "type": self.type,
            "title": self.title,
            "message": self.message,
            "link": self.link,
            "is_read": self.is_read,
            "created_at": self.created_at.strftime('%Y-%m-%d %H:%M:%S') if self.created_at else "",
        }


def create_notification(recipient_email, recipient_role, ntype, title, message, link=None):
    """
    Best-effort notification creation. Failures here (e.g. a transient DB
    hiccup) must never break the action that triggered them — submitting a
    complaint or updating its status is the important part; the
    notification is a courtesy on top of it.
    """
    try:
        note = Notification(
            recipient_email=recipient_email,
            recipient_role=recipient_role,
            type=ntype,
            title=title[:150],
            message=message[:500],
            link=link,
        )
        db.session.add(note)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        print(f"[HOSTEL APP] Failed to create notification for {recipient_email}: {e}")


class ActivityLog(db.Model):
    """Tracks events across the system (complaint lifecycle, notices,
    warden updates) so a future Recent Activity Timeline can be built from
    real history instead of approximating from created/updated timestamps.
    This table only starts capturing data going forward — it has no
    knowledge of anything that happened before it was added."""
    __tablename__ = "activity_log"
    id = db.Column(db.Integer, primary_key=True)
    event_type = db.Column(db.String(40), nullable=False, index=True)
    # e.g. 'complaint_created', 'status_updated', 'resolved', 'complaint_deleted',
    # 'notice_sent', 'warden_updated'
    complaint_id = db.Column(db.Integer, nullable=True, index=True)
    hostel_type = db.Column(db.String(10), nullable=True, index=True)
    block = db.Column(db.String(5), nullable=True, index=True)
    actor = db.Column(db.String(120), nullable=True)
    message = db.Column(db.String(300), nullable=False)
    created_at = db.Column(db.DateTime, nullable=False, default=datetime.utcnow, index=True)

    def to_dict(self):
        return {
            "id": self.id,
            "event_type": self.event_type,
            "complaint_id": self.complaint_id,
            "hostel_type": self.hostel_type,
            "block": self.block,
            "actor": self.actor,
            "message": self.message,
            "created_at": self.created_at.strftime('%Y-%m-%d %H:%M:%S') if self.created_at else "",
        }


def log_activity(event_type, message, complaint=None, hostel_type=None, block=None, actor=None):
    """Best-effort activity log write — same non-blocking philosophy as
    create_notification: never let logging break the real action."""
    try:
        entry = ActivityLog(
            event_type=event_type,
            message=message[:300],
            complaint_id=(complaint.id if complaint else None),
            hostel_type=(hostel_type if hostel_type is not None else (complaint.hostel_type if complaint else None)),
            block=(block if block is not None else (complaint.block if complaint else None)),
            actor=actor or session.get("user"),
        )
        db.session.add(entry)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        print(f"[HOSTEL APP] Failed to log activity ({event_type}): {e}")


def _allowed_image(filename):
    if not filename or "." not in filename:
        return False
    ext = filename.rsplit(".", 1)[1].lower()
    return ext in {"png", "jpg", "jpeg", "webp", "gif"}
 
 
def _tables_missing():
    """
    --- DB INIT FIX ---
    Returns the set of expected tables that do NOT yet exist in the
    connected database. This is what lets us skip db.create_all()
    entirely on every warm/normal request — we only ever attempt to
    create tables when one is actually missing (e.g. the very first
    time the app talks to a brand-new Neon database).
    """
    required_tables = {"complaints", "issue_counter", "student_profiles", "wardens", "notices", "notifications"}
    try:
        inspector = sa_inspect(db.engine)
        existing = set(inspector.get_table_names())
    except Exception as e:
        # If we can't even inspect the DB, treat as "unknown" and let
        # the create_all() call below (which is itself guarded) decide.
        print(f"[HOSTEL APP] Could not inspect database schema: {e}")
        return required_tables
    return required_tables - existing
 
 
def _init_db():
    global DB_READY
    if not DB_READY:
        db_type = "PostgreSQL" if os.environ.get("DATABASE_URL") else "SQLite"
 
        # --- DB INIT FIX (root cause of the UniqueViolation) ---
        # The old code called db.create_all() unconditionally on every
        # request. On Vercel each request/cold-start is a fresh process,
        # so DB_READY resets to False constantly, and db.create_all() kept
        # firing again and again against the SAME already-initialized Neon
        # database. Under concurrent cold starts, two invocations could
        # both decide the "complaints" sequence didn't exist yet and both
        # try to create it at the same instant, producing:
        #   duplicate key value violates unique constraint "pg_class_relname_nsp_index"
        #
        # Fix: check with the inspector first. If every expected table is
        # already there (the normal case in production), skip create_all()
        # completely — no DDL, no race, no error, and it's fast.
        missing = _tables_missing()
        if missing:
            print(f"[HOSTEL APP] Initializing with {db_type} database... missing tables: {sorted(missing)}")
            try:
                db.create_all()
                db.session.commit()
            except Exception as e:
                # Belt-and-braces: if a concurrent cold start won the race
                # and created the tables/sequences a split second before us,
                # Postgres will raise a duplicate/already-exists error here.
                # That's not a real failure — the schema ended up correct
                # either way — so we roll back and continue instead of
                # crashing the request.
                db.session.rollback()
                print(f"[HOSTEL APP] create_all() raced with a concurrent init "
                      f"(safe to ignore, schema already exists): {e}")
        else:
            print(f"[HOSTEL APP] {db_type} schema already present, skipping create_all().")
 
        columns_to_add = [
            ("password", "VARCHAR(255)"),
            ("hostel_type", "VARCHAR(10)"),
            ("block", "VARCHAR(10)"),
            ("floor", "VARCHAR(10)"),
            ("room_no", "VARCHAR(20)")
        ]
        for col_name, col_type in columns_to_add:
            try:
                db.session.execute(db.text(f"SELECT {col_name} FROM student_profiles LIMIT 1"))
            except Exception:
                db.session.rollback()
                print(f"[HOSTEL APP] Adding column '{col_name}' to 'student_profiles'...")
                try:
                    db.session.execute(db.text(f"ALTER TABLE student_profiles ADD COLUMN {col_name} {col_type}"))
                    db.session.commit()
                except Exception as e:
                    print(f"[HOSTEL APP] Error adding column '{col_name}': {e}")
                    db.session.rollback()
 
        # Add floor to complaints table if not present
        try:
            db.session.execute(db.text("SELECT floor FROM complaints LIMIT 1"))
        except Exception:
            db.session.rollback()
            print("[HOSTEL APP] Adding column 'floor' to 'complaints'...")
            try:
                db.session.execute(db.text("ALTER TABLE complaints ADD COLUMN floor VARCHAR(10)"))
                db.session.commit()
            except Exception as e:
                print(f"[HOSTEL APP] Error adding column 'floor' to 'complaints': {e}")
                db.session.rollback()
 
        # Add resolved_by to notices table if not present
        try:
            db.session.execute(db.text("SELECT resolved_by FROM notices LIMIT 1"))
        except Exception:
            db.session.rollback()
            print("[HOSTEL APP] Adding column 'resolved_by' to 'notices'...")
            try:
                db.session.execute(db.text("ALTER TABLE notices ADD COLUMN resolved_by TIMESTAMP"))
                db.session.commit()
            except Exception as e:
                print(f"[HOSTEL APP] Error adding column 'resolved_by' to 'notices': {e}")
                db.session.rollback()
 
        # Add indexes for columns that are now filtered/sorted on directly
        # (search, status/category/priority filters, date range, sort) —
        # matters once the complaints table has any real volume. Safe to
        # run every startup: IF NOT EXISTS makes it a no-op after the first
        # time, on both SQLite and Postgres.
        indexes_to_add = [
            ("ix_complaints_status", "complaints", "status"),
            ("ix_complaints_created_at", "complaints", "created_at"),
            ("ix_complaints_category", "complaints", "category"),
            ("ix_complaints_priority", "complaints", "priority"),
            ("ix_complaints_hostel_type", "complaints", "hostel_type"),
            ("ix_complaints_block", "complaints", "block"),
            # Added for the scoped dashboards/analytics/reports: every one of
            # them filters on hostel_type+block together, and the new
            # resolved-today / monthly-resolution / avg-resolution-time
            # analytics all range-filter on updated_at.
            ("ix_complaints_updated_at", "complaints", "updated_at"),
            ("ix_complaints_hostel_block", "complaints", "hostel_type, block"),
        ]
        for idx_name, table_name, col_name in indexes_to_add:
            try:
                db.session.execute(db.text(f"CREATE INDEX IF NOT EXISTS {idx_name} ON {table_name} ({col_name})"))
                db.session.commit()
            except Exception as e:
                print(f"[HOSTEL APP] Could not ensure index '{idx_name}': {e}")
                db.session.rollback()
 
        # Initialize default wardens if table is empty
        try:
            db.session.execute(db.text("SELECT 1 FROM wardens LIMIT 1"))
        except Exception:
            db.session.rollback()
            print("[HOSTEL APP] Table 'wardens' might not exist or need creation...")
            
        try:
            if Warden.query.count() == 0:
                default_wardens = [
                    # Boys Hostel
                    Warden(hostel_type="boys", block="A", name="Mr. Rajesh Kumar", contact="+91 98765 43210"),
                    Warden(hostel_type="boys", block="B", name="Mr. Suresh Raina", contact="+91 98765 43211"),
                    Warden(hostel_type="boys", block="C", name="Mr. Amit Sharma", contact="+91 98765 43212"),
                    Warden(hostel_type="boys", block="D", name="Mr. Vijay Singh", contact="+91 98765 43213"),
                    Warden(hostel_type="boys", block="E", name="Mr. Dinesh Karthik", contact="+91 98765 43214"),
                    Warden(hostel_type="boys", block="F", name="Mr. Ramesh Sen", contact="+91 98765 43215"),
                    # Girls Hostel
                    Warden(hostel_type="girls", block="A", name="Mrs. Priya Patel", contact="+91 98765 43216"),
                    Warden(hostel_type="girls", block="B", name="Mrs. Lakshmi Roy", contact="+91 98765 43217"),
                    Warden(hostel_type="girls", block="C", name="Mrs. Sunita Rao", contact="+91 98765 43218"),
                    Warden(hostel_type="girls", block="D", name="Mrs. Anita Desai", contact="+91 98765 43219"),
                    Warden(hostel_type="girls", block="E", name="Mrs. Radha Krishnan", contact="+91 98765 43220"),
                    Warden(hostel_type="girls", block="F", name="Mrs. Deepa Nair", contact="+91 98765 43221"),
                ]
                db.session.bulk_save_objects(default_wardens)
                db.session.commit()
                print("[HOSTEL APP] Default wardens initialized.")
        except Exception as e:
            print(f"[HOSTEL APP] Error initializing wardens: {e}")
            db.session.rollback()
 
        # Initialize default notices if table is empty
        try:
            if Notice.query.count() == 0:
                default_notices = [
                    Notice(hostel_type="all", block="B", title="Water Interruption", content="Block B — 6 AM–9 AM every Sunday."),
                    Notice(hostel_type="all", block="all", title="Wi-Fi Upgrade", content="New routers on floors 3 & 4."),
                    Notice(hostel_type="all", block="all", title="Response SLA", content="All issues resolved within 24 hrs."),
                ]
                db.session.bulk_save_objects(default_notices)
                db.session.commit()
                print("[HOSTEL APP] Default notices initialized.")
        except Exception as e:
            print(f"[HOSTEL APP] Error initializing notices: {e}")
            db.session.rollback()
 
        counter = IssueCounter.get()
        if counter.total == 0:
            existing = Complaint.query.count()
            if existing:
                counter.total = existing
                db.session.commit()
        print(f"[HOSTEL APP] Database ready. Total complaints: {IssueCounter.get().total}")
        DB_READY = True
 
 
@app.route("/uploads/<path:filename>")
def uploaded_file(filename):
    return send_from_directory(UPLOAD_DIR, filename)


# --- CSRF protection -----------------------------------------------------
# Every session gets a random token the moment a page is rendered; every
# state-changing request (POST/PUT/PATCH/DELETE) must echo it back, either
# as a form field named csrf_token or an X-CSRF-Token header. This is a
# deliberately dependency-free implementation (no Flask-WTF) so it doesn't
# touch unrelated app config, but it protects the same set of forms.
def _get_csrf_token():
    token = session.get("_csrf_token")
    if not token:
        token = secrets.token_hex(32)
        session["_csrf_token"] = token
    return token


@app.context_processor
def inject_csrf_token():
    return dict(csrf_token=_get_csrf_token)


@app.before_request
def _csrf_protect():
    if request.method in ("POST", "PUT", "PATCH", "DELETE"):
        session_token = session.get("_csrf_token")
        submitted_token = request.form.get("csrf_token") or request.headers.get("X-CSRF-Token")
        if not session_token or not submitted_token or not hmac.compare_digest(session_token, submitted_token):
            abort(400, description="Your session expired or the form was tampered with. Please refresh and try again.")


def _is_valid_image_upload(file_storage):
    """
    Confirms the uploaded file is actually a decodable image of an allowed
    type — not just something with a trusted-looking file extension. This
    closes the gap where a renamed script or polyglot file could otherwise
    pass the extension-only check.
    """
    try:
        from PIL import Image
        file_storage.stream.seek(0)
        img = Image.open(file_storage.stream)
        img.verify()
        file_storage.stream.seek(0)
        return (img.format or "").upper() in {"PNG", "JPEG", "JPG", "WEBP", "GIF"}
    except Exception:
        return False
    finally:
        try:
            file_storage.stream.seek(0)
        except Exception:
            pass


@app.before_request
def _setup():
    # --- DB INIT FIX: _init_db() no longer runs here. Running it on every
    # single request was the root cause of the UniqueViolation — it doesn't
    # need to happen per-request at all, only once when the app/process
    # starts. See the call to _init_db() further below (module scope) for
    # where it now actually happens.
    try:
        os.makedirs(UPLOAD_DIR, exist_ok=True)
    except OSError:
        pass
 
 
# --- DB INIT FIX: Run database initialization ONCE, at import time
# (i.e. once per cold start on Vercel, never per-request), instead of
# inside @app.before_request. _init_db() itself is still safe to call
# more than once (it checks DB_READY and inspects existing tables), but
# this placement means normal warm requests never touch this code path
# at all — only a fresh process/cold start does.
#
# Schema changes going forward are handled by the self-healing column
# checks below (SELECT ... LIMIT 1 -> ALTER TABLE ADD COLUMN if missing),
# combined with the inspector-guarded create_all() as a safety net for
# first-time/empty databases. No external migration tool is required.
try:
    with app.app_context():
        _init_db()
except Exception as e:
    # Never let a DB hiccup at cold-start prevent the app from booting —
    # individual routes already handle DB errors (see error handlers and
    # try/except blocks throughout). Log and continue; _init_db() will
    # simply be retried (DB_READY is still False) on the next cold start.
    print(f"[HOSTEL APP] Deferred: database initialization failed at startup: {e}")
 
@app.after_request
def add_security_headers(response):
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["X-XSS-Protection"] = "1; mode=block"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    return response
 
 
@app.route("/")
def login():
    if "user" in session:
        return redirect(url_for("welcome"))
    return render_template("login.html")
 
 
def _start_authenticated_session(email, role, full_name):
    # Clear first to avoid session fixation: a token issued before login
    # (e.g. to an attacker who tricked the victim into using their session
    # id) must not carry authenticated privileges forward.
    session.clear()
    session.permanent = True
    session["user"] = email
    session["role"] = role
    session["full_name"] = full_name
    session["login_time"] = datetime.utcnow().isoformat()


@app.route("/login", methods=["POST"])
def handle_login():
    email = (request.form.get("email") or "").strip().lower()
    password = (request.form.get("password") or "").strip()
 
    if not email or not password:
        return render_template("login.html", error="Please enter email and password.")
 
    email = sanitize_string(email, 120)
    
    if not re.match(r'^[\w\.-]+@[\w\.-]+\.\w+$', email):
        return render_template("login.html", error="Invalid email format.")
 
    attempt_key = _login_attempt_key(email)
    if _too_many_login_attempts(attempt_key):
        return render_template(
            "login.html",
            error="Too many failed login attempts. Please wait a few minutes and try again."
        )
 
    if email == ADMIN_EMAIL and hmac.compare_digest(password, ADMIN_PASSWORD):
        _clear_login_attempts(attempt_key)
        _start_authenticated_session(email, "admin", "Administrator")
        return redirect(url_for("welcome"))
 
    # Check DB profile next
    profile = StudentProfile.query.filter_by(email=email).first()
    if profile:
        if verify_and_maybe_upgrade_password(profile, password):
            _clear_login_attempts(attempt_key)
            _start_authenticated_session(email, "student", profile.name)
            return redirect(url_for("welcome"))
        _record_failed_login(attempt_key)
        return render_template("login.html", error="Incorrect password. Please try again.")
 
    _record_failed_login(attempt_key)
    return render_template("login.html", error="Email not registered. Use your official SECE email.")
 
 
@app.route("/login/google")
def google_login():
    if "user" in session:
        return redirect(url_for("welcome"))
    redirect_uri = url_for("google_authorize", _external=True)
    return google_oauth.authorize_redirect(redirect_uri)
 
 
@app.route("/login/google/callback")
def google_authorize():
    try:
        token = google_oauth.authorize_access_token()
    except Exception:
        flash("Google sign-in failed or was cancelled. Please try again.", "error")
        return redirect(url_for("login"))
 
    userinfo = token.get("userinfo") or {}
    if not userinfo:
        try:
            userinfo = google_oauth.parse_id_token(token, nonce=session.get("oauth_nonce")) or {}
        except Exception:
            userinfo = {}
 
    email = (userinfo.get("email") or "").strip().lower()
    full_name = sanitize_string(userinfo.get("name") or "", 120)
    picture = userinfo.get("picture") or ""
    if not email or not email.endswith("@sece.ac.in"):
        session.clear()
        flash("Only SECE email accounts are allowed to access this website.", "error")
        return redirect(url_for("login"))
 
    # Valid SECE account: establish a fresh, secure session.
    session.clear()
    session.permanent = True
    session["user"] = email
    session["full_name"] = full_name or email.split("@")[0]
    session["picture"] = picture
    session["login_time"] = datetime.utcnow().isoformat()
    session["role"] = "admin" if email == ADMIN_EMAIL else "student"
    session["auth_provider"] = "google"
    return redirect(url_for("welcome"))
 
 
@app.route("/signup", methods=["POST"])
def handle_signup():
    name = sanitize_string(request.form.get("name", ""), 120).strip()
    roll_number = sanitize_string(request.form.get("roll_number", ""), 50).strip()
    dept = sanitize_string(request.form.get("dept", ""), 100).strip()
    year = sanitize_string(request.form.get("year", ""), 10).strip()
    phone = sanitize_string(request.form.get("phone", ""), 20).strip()
    email = sanitize_string(request.form.get("email", ""), 120).lower().strip()
    password = (request.form.get("password") or "").strip()
    hostel_type = sanitize_string(request.form.get("hostel_type", ""), 10).strip()
    block = sanitize_string(request.form.get("block", ""), 10).strip()
    floor = sanitize_string(request.form.get("floor", ""), 10).strip()
    room_no = sanitize_string(request.form.get("room_no", ""), 20).strip()
 
    if not (name and roll_number and dept and year and phone and email and password and hostel_type and block and floor and room_no):
        return render_template("login.html", error="All fields are required.", show_signup=True)
 
    if not email.endswith("@sece.ac.in"):
        return render_template("login.html", error="Use your college official email", show_signup=True)
 
    # Validate phone number (must be 10 digits)
    if not re.match(r'^\d{10}$', phone):
        return render_template("login.html", error="Phone number must be a 10-digit number.", show_signup=True)
 
    if len(password) < 8:
        return render_template("login.html", error="Password must be at least 8 characters long.", show_signup=True)
 
    # Check if user already exists
    existing = StudentProfile.query.filter_by(email=email).first()
    if existing:
        return render_template("login.html", error="Email is already registered. Please log in.", show_signup=False)
 
    try:
        new_profile = StudentProfile(
            email=email,
            name=name,
            roll_number=roll_number,
            dept=dept,
            year=year,
            phone=phone,
            password=hash_password(password),
            hostel_type=hostel_type,
            block=block,
            floor=floor,
            room_no=room_no
        )
        db.session.add(new_profile)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        print(f"[HOSTEL APP] Signup failed for {email}: {e}")
        return render_template("login.html", error="Registration failed. Please try again.", show_signup=True)
 
    return render_template("login.html", success="Registration successful! Please log in with your credentials.", show_signup=False)
 
 
HOSTEL_LABELS = {"boys": "Boys Hostel", "girls": "Girls Hostel"}


def dashboard_title_for(hostel, block):
    """Returns (title, description) for the current admin dashboard scope."""
    if hostel == "all":
        return "All Hostels", "Manage complaints across every hostel and block."
    hostel_label = HOSTEL_LABELS.get(hostel, hostel.title())
    if block == "all":
        return hostel_label, f"Manage all complaints raised inside {hostel_label}."
    return f"{hostel_label} — Block {block}", f"Manage complaints raised in {hostel_label} Block {block}."


def compute_scope_stats(hostel, block):
    """Computes all top-of-dashboard stat card numbers scoped to a given
    hostel/block selection ('all' means no filter on that dimension)."""
    q = Complaint.query
    if hostel != "all":
        q = q.filter(db.func.lower(Complaint.hostel_type) == hostel)
    if block != "all":
        q = q.filter(db.func.upper(Complaint.block) == block)

    status_counts = dict(q.with_entities(Complaint.status, db.func.count(Complaint.id)).group_by(Complaint.status).all())
    total = sum(status_counts.values())
    pending = status_counts.get("Pending", 0)
    in_progress = status_counts.get("In Progress", 0)
    resolved = status_counts.get("Resolved", 0)

    pri_counts = dict(q.with_entities(Complaint.priority, db.func.count(Complaint.id)).group_by(Complaint.priority).all())
    high = pri_counts.get("High", 0)
    medium = pri_counts.get("Moderate", 0)
    low = pri_counts.get("Low", 0)

    now = datetime.utcnow()
    today_start = datetime(now.year, now.month, now.day)
    today_count = q.filter(Complaint.created_at >= today_start).count()
    resolved_today = q.filter(Complaint.status == "Resolved", Complaint.updated_at >= today_start).count()

    resolved_rows = q.filter(Complaint.status == "Resolved").with_entities(Complaint.created_at, Complaint.updated_at).all()
    durations = [(u - c).total_seconds() for c, u in resolved_rows if u and c]
    avg_resolution_hours = round((sum(durations) / len(durations)) / 3600, 1) if durations else 0

    this_month_start = datetime(now.year, now.month, 1)
    last_month_end = this_month_start - timedelta(seconds=1)
    last_month_start = datetime(last_month_end.year, last_month_end.month, 1)
    this_month_count = q.filter(Complaint.created_at >= this_month_start).count()
    last_month_count = q.filter(Complaint.created_at >= last_month_start, Complaint.created_at < this_month_start).count()
    if last_month_count > 0:
        monthly_growth = round(((this_month_count - last_month_count) / last_month_count) * 100, 1)
    else:
        monthly_growth = 100.0 if this_month_count > 0 else 0.0

    return dict(
        total=total, active=total, pending=pending, in_progress=in_progress, resolved=resolved,
        high=high, medium=medium, low=low,
        today=today_count, resolved_today=resolved_today,
        avg_resolution_hours=avg_resolution_hours, monthly_growth=monthly_growth,
    )


WEEKDAY_LABELS = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
HOUR_BUCKETS = [("Night", 0, 6), ("Morning", 6, 12), ("Afternoon", 12, 18), ("Evening", 18, 24)]


def _scope_query(hostel, block):
    q = Complaint.query
    if hostel != "all":
        q = q.filter(db.func.lower(Complaint.hostel_type) == hostel)
    if block != "all":
        q = q.filter(db.func.upper(Complaint.block) == block)
    return q


def compute_extended_analytics(hostel, block):
    """Powers the additional analytics charts (heatmap, weekly trend, top
    wardens, frequent locations, monthly resolution comparison, avg
    resolution trend, block comparison) plus the extended JSON API. Kept
    separate from compute_scope_stats (the top stat cards) since these are
    heavier, chart-shaped queries that not every caller needs."""
    q = _scope_query(hostel, block)
    now = datetime.utcnow()
    from sqlalchemy import func as _func

    # --- Weekly trend: daily complaint counts for the last 7 days ---
    # One bounded query + Python-side bucketing instead of 7 separate
    # per-day COUNT queries.
    week_start = datetime(now.year, now.month, now.day) - timedelta(days=6)
    week_rows = q.filter(Complaint.created_at >= week_start).with_entities(Complaint.created_at).all()
    day_buckets = {}
    for i in range(7):
        day_buckets[(week_start + timedelta(days=i)).date()] = 0
    for (created_at,) in week_rows:
        if created_at and created_at.date() in day_buckets:
            day_buckets[created_at.date()] += 1
    weekly_trend = [{"label": day.strftime("%a %d"), "count": count} for day, count in sorted(day_buckets.items())]

    # --- Heatmap: weekday x time-of-day bucket, last 60 days ---
    heatmap_rows = q.filter(Complaint.created_at >= now - timedelta(days=60)).with_entities(Complaint.created_at).all()
    heatmap = {wd: {bucket[0]: 0 for bucket in HOUR_BUCKETS} for wd in WEEKDAY_LABELS}
    for (created_at,) in heatmap_rows:
        if not created_at:
            continue
        wd = WEEKDAY_LABELS[created_at.weekday()]
        hr = created_at.hour
        for label, start, end in HOUR_BUCKETS:
            if start <= hr < end:
                heatmap[wd][label] += 1
                break
    heatmap_data = [{"day": wd, **heatmap[wd]} for wd in WEEKDAY_LABELS]

    # --- Top wardens: resolved-complaint count within each warden's own block ---
    # Two grouped queries (total per hostel/block, resolved per hostel/block)
    # instead of 2 queries PER warden.
    wardens_scope = Warden.query.all()
    if hostel != "all":
        wardens_scope = [w for w in wardens_scope if w.hostel_type == hostel]
    if block != "all":
        wardens_scope = [w for w in wardens_scope if w.block == block]

    totals_by_scope = {
        (h, b): n for h, b, n in Complaint.query.with_entities(
            Complaint.hostel_type, Complaint.block, _func.count(Complaint.id)
        ).group_by(Complaint.hostel_type, Complaint.block).all()
    }
    resolved_by_scope = {
        (h, b): n for h, b, n in Complaint.query.filter(Complaint.status == "Resolved").with_entities(
            Complaint.hostel_type, Complaint.block, _func.count(Complaint.id)
        ).group_by(Complaint.hostel_type, Complaint.block).all()
    }
    top_wardens = []
    for w in wardens_scope:
        key = (w.hostel_type, w.block)
        # totals_by_scope keys use the DB's exact casing; try a couple of
        # sensible variants rather than a case-insensitive query per warden.
        total_count = totals_by_scope.get(key) or totals_by_scope.get((w.hostel_type.lower(), w.block.upper())) or 0
        resolved_count = resolved_by_scope.get(key) or resolved_by_scope.get((w.hostel_type.lower(), w.block.upper())) or 0
        top_wardens.append({
            "name": w.name, "hostel": w.hostel_type, "block": w.block,
            "resolved": resolved_count, "total": total_count,
            "resolution_rate": round((resolved_count / total_count) * 100, 1) if total_count else 0,
        })
    top_wardens.sort(key=lambda x: x["resolved"], reverse=True)
    top_wardens = top_wardens[:5]

    # --- Most frequent complaint locations (room numbers) ---
    room_rows = (
        q.with_entities(Complaint.room_number, _func.count(Complaint.id))
        .group_by(Complaint.room_number)
        .order_by(_func.count(Complaint.id).desc())
        .limit(5).all()
    )
    frequent_locations = [{"room": r or "—", "count": n} for r, n in room_rows]

    # --- Monthly resolution comparison + avg resolution time trend, 6 months ---
    # Two bounded queries (created-in-range, resolved-in-range) + Python
    # bucketing, instead of 12 separate per-month COUNT/fetch queries.
    six_months_start = datetime(now.year, now.month, 1)
    for _ in range(5):
        six_months_start = (six_months_start - timedelta(days=1)).replace(day=1)

    month_keys = []
    for i in range(5, -1, -1):
        mm = now.month - i
        yy = now.year
        while mm <= 0:
            mm += 12
            yy -= 1
        month_keys.append((yy, mm))

    created_counts = {k: 0 for k in month_keys}
    created_rows = q.filter(Complaint.created_at >= six_months_start).with_entities(Complaint.created_at).all()
    for (created_at,) in created_rows:
        if created_at:
            key = (created_at.year, created_at.month)
            if key in created_counts:
                created_counts[key] += 1

    resolved_counts = {k: 0 for k in month_keys}
    resolved_durations = {k: [] for k in month_keys}
    resolved_rows = q.filter(
        Complaint.status == "Resolved", Complaint.updated_at >= six_months_start
    ).with_entities(Complaint.created_at, Complaint.updated_at).all()
    for created_at, updated_at in resolved_rows:
        if not updated_at:
            continue
        key = (updated_at.year, updated_at.month)
        if key in resolved_counts:
            resolved_counts[key] += 1
            if created_at:
                resolved_durations[key].append((updated_at - created_at).total_seconds())

    monthly_resolution = []
    avg_resolution_trend = []
    for (yy, mm) in month_keys:
        label = datetime(yy, mm, 1).strftime("%b %Y")
        monthly_resolution.append({"label": label, "created": created_counts[(yy, mm)], "resolved": resolved_counts[(yy, mm)]})
        durations = resolved_durations[(yy, mm)]
        avg_hours = round((sum(durations) / len(durations)) / 3600, 1) if durations else 0
        avg_resolution_trend.append({"label": label, "avg_hours": avg_hours})

    # --- Block comparison: siblings at the current scope level ---
    block_comparison = []
    if hostel == "all":
        for h in sorted(VALID_HOSTEL_TYPES):
            hq = Complaint.query.filter(db.func.lower(Complaint.hostel_type) == h)
            block_comparison.append({
                "label": HOSTEL_LABELS.get(h, h.title()),
                "total": hq.count(), "resolved": hq.filter(Complaint.status == "Resolved").count(),
            })
    else:
        for b in sorted(VALID_BLOCKS):
            bq = Complaint.query.filter(db.func.lower(Complaint.hostel_type) == hostel, db.func.upper(Complaint.block) == b)
            total_b = bq.count()
            if total_b == 0:
                continue
            block_comparison.append({
                "label": f"Block {b}", "total": total_b, "resolved": bq.filter(Complaint.status == "Resolved").count(),
            })

    return dict(
        weekly_trend=weekly_trend,
        heatmap_data=heatmap_data,
        top_wardens=top_wardens,
        frequent_locations=frequent_locations,
        monthly_resolution=monthly_resolution,
        avg_resolution_trend=avg_resolution_trend,
        block_comparison=block_comparison,
    )


def welcome_view(forced_hostel=None, forced_block=None):
    """Renders the admin/student dashboard. When forced_hostel/forced_block
    are given (from a dedicated /admin/... route), the dashboard is locked
    to that scope; otherwise the scope comes from ?hostel=&block= query
    params on /welcome, preserving old links/bookmarks."""
    wardens = Warden.query.all()
    notices = []
    dashboard_scope = None
 
    if session.get("role") == "admin":
        complaints = Complaint.query.order_by(Complaint.created_at.desc()).all()
 
        from sqlalchemy import func
        cat_rows = db.session.query(Complaint.category, func.count(Complaint.id)).group_by(Complaint.category).all()
        cat_labels = [r[0] for r in cat_rows]
        cat_counts = [r[1] for r in cat_rows]
 
        pri_rows = db.session.query(Complaint.priority, func.count(Complaint.id)).group_by(Complaint.priority).all()
        pri_dict = {r[0]: r[1] for r in pri_rows}
 
        # Block-wise complaint data for admin
        block_rows = db.session.query(Complaint.block, func.count(Complaint.id)).filter(Complaint.block != None).group_by(Complaint.block).all()
        block_data = {r[0]: r[1] for r in block_rows}
        # Also get block+status breakdown
        block_status_rows = db.session.query(Complaint.block, Complaint.status, func.count(Complaint.id)).filter(Complaint.block != None).group_by(Complaint.block, Complaint.status).all()
        block_status_data = {}
        for blk, sts, cnt in block_status_rows:
            if blk not in block_status_data:
                block_status_data[blk] = {}
            block_status_data[blk][sts] = cnt
 
        # Hostel-type (Boys / Girls) split for bar chart
        hostel_rows = db.session.query(
            Complaint.hostel_type, Complaint.category, func.count(Complaint.id)
        ).filter(Complaint.hostel_type != None).group_by(Complaint.hostel_type, Complaint.category).all()
        hostel_category_data = {"boys": {}, "girls": {}}
        for ht, cat, cnt in hostel_rows:
            if ht in hostel_category_data:
                hostel_category_data[ht][cat] = cnt
 
        # Simple totals per hostel for the summary bar chart
        hostel_total_rows = db.session.query(Complaint.hostel_type, func.count(Complaint.id)).filter(Complaint.hostel_type != None).group_by(Complaint.hostel_type).all()
        hostel_totals = {r[0]: r[1] for r in hostel_total_rows}
 
        # Hostel status breakdown
        hostel_status_rows = db.session.query(Complaint.hostel_type, Complaint.status, func.count(Complaint.id)).filter(Complaint.hostel_type != None).group_by(Complaint.hostel_type, Complaint.status).all()
        hostel_status_data = {"boys": {}, "girls": {}}
        for ht, sts, cnt in hostel_status_rows:
            if ht in hostel_status_data:
                hostel_status_data[ht][sts] = cnt
 
        # Hostel + Block cross data: {hostel_type: {block: {status: count}}}
        hostel_block_rows = db.session.query(
            Complaint.hostel_type, Complaint.block, Complaint.status, func.count(Complaint.id)
        ).filter(Complaint.hostel_type != None, Complaint.block != None).group_by(
            Complaint.hostel_type, Complaint.block, Complaint.status
        ).all()
        hostel_block_data = {"boys": {}, "girls": {}}
        for ht, blk, sts, cnt in hostel_block_rows:
            if ht in hostel_block_data:
                if blk not in hostel_block_data[ht]:
                    hostel_block_data[ht][blk] = {}
                hostel_block_data[ht][blk][sts] = cnt
 
        now = datetime.utcnow()
        month_labels, monthly_issued, monthly_resolved = [], [], []
        for i in range(5, -1, -1):
            m = now.month - i
            y = now.year
            while m <= 0:
                m += 12
                y -= 1
            month_labels.append(datetime(y, m, 1).strftime("%b %Y"))
            issued = sum(1 for c in complaints if c.created_at.month == m and c.created_at.year == y)
            resolved = sum(1 for c in complaints if c.status == "Resolved" and c.updated_at.month == m and c.updated_at.year == y)
            monthly_issued.append(issued)
            monthly_resolved.append(resolved)
    else:
        complaints = Complaint.query.filter_by(student_email=session["user"]).order_by(Complaint.created_at.desc()).all()
        cat_labels = cat_counts = month_labels = monthly_issued = monthly_resolved = []
        pri_dict = {}
        block_data = {}
        block_status_data = {}
        hostel_category_data = {}
        hostel_totals = {}
        hostel_status_data = {}
        hostel_block_data = {}
 
    # --- Real DB-level pagination for the visible complaint table ---
    # (The charts/complaints_json above intentionally still run against the
    # full dataset — they power the existing block/hostel drill-down
    # explorer, which reads from JSON, not from scanning the table's DOM
    # rows, so pagination doesn't affect them.)
    base_table_query = Complaint.query if session.get("role") == "admin" else Complaint.query.filter_by(student_email=session["user"])

    if session.get("role") == "admin":
        # Scope (hostel/block) comes from the dedicated route if this view
        # was reached via /admin/boys, /admin/girls/block-c, etc.; otherwise
        # it falls back to ?hostel=&block= query params on /welcome.
        effective_hostel = (forced_hostel or request.args.get("hostel") or "all").lower()
        effective_block = (forced_block or request.args.get("block") or "all").upper()
        if effective_hostel not in VALID_HOSTEL_TYPES:
            effective_hostel = "all"
        if effective_block not in VALID_BLOCKS:
            effective_block = "all"

        scope_stats = compute_scope_stats(effective_hostel, effective_block)
        if effective_hostel == "all" and effective_block == "all":
            # Preserve the historical lifetime counter (survives deletions)
            # on the top-level "All Hostels" view only.
            scope_stats["total"] = IssueCounter.get().total
        stats = scope_stats
        extended_analytics = compute_extended_analytics(effective_hostel, effective_block)

        activity_log_query = ActivityLog.query
        if effective_hostel != "all":
            activity_log_query = activity_log_query.filter(db.func.lower(ActivityLog.hostel_type) == effective_hostel)
        if effective_block != "all":
            activity_log_query = activity_log_query.filter(db.func.upper(ActivityLog.block) == effective_block)
        recent_activity = activity_log_query.order_by(ActivityLog.created_at.desc()).limit(15).all()

        title, desc = dashboard_title_for(effective_hostel, effective_block)
        dashboard_scope = dict(hostel=effective_hostel, block=effective_block, title=title, desc=desc)

        # Build the filter args used for the table query and for the filter
        # UI, with hostel/block forced to the current scope.
        filter_args = request.args.to_dict()
        filter_args["hostel"] = effective_hostel
        filter_args["block"] = effective_block
    else:
        extended_analytics = None
        recent_activity = []
        status_counts = dict(
            base_table_query.with_entities(Complaint.status, db.func.count(Complaint.id)).group_by(Complaint.status).all()
        )
        active = sum(status_counts.values())
        stats = dict(
            total=active, active=active,
            pending=status_counts.get("Pending", 0),
            in_progress=status_counts.get("In Progress", 0),
            resolved=status_counts.get("Resolved", 0),
        )
        filter_args = request.args.to_dict()

    filtered_table_query = apply_complaint_sort(
        apply_complaint_filters(base_table_query, filter_args),
        (filter_args.get("sort") or "latest").strip()
    )
    page = request.args.get("page", 1, type=int)
    if not page or page < 1:
        page = 1
    complaints_page = filtered_table_query.paginate(page=page, per_page=COMPLAINTS_PAGE_SIZE, error_out=False)

    if session.get("role") == "admin" and complaints_page.items:
        # Enrich each visible row with the student's roll number and the
        # warden assigned to that complaint's hostel/block — both derived
        # from existing tables, no schema changes needed.
        page_emails = {c.student_email for c in complaints_page.items}
        profiles_by_email = {
            p.email: p for p in StudentProfile.query.filter(StudentProfile.email.in_(page_emails)).all()
        }
        warden_by_scope = {(w.hostel_type, w.block): w for w in wardens}
        for c in complaints_page.items:
            prof = profiles_by_email.get(c.student_email)
            c.roll_number_display = (prof.roll_number if prof and prof.roll_number else "—")
            w = warden_by_scope.get(((c.hostel_type or "").lower(), (c.block or "").upper()))
            c.assigned_warden_display = w.name if w else "Unassigned"

    # Build page-aware pagination/reset URLs so they respect whichever
    # route (/welcome, /admin/boys, /admin/girls/block-c, ...) is current.
    from urllib.parse import urlencode
    _qs_keys = ["q", "status", "category", "priority", "date_from", "date_to", "sort", "student_name", "room", "warden"]
    if forced_hostel is None:
        _qs_keys += ["hostel", "block"]

    def _page_url(target_page):
        qs = {k: filter_args.get(k) for k in _qs_keys if filter_args.get(k) not in (None, "", "all")}
        qs["page"] = target_page
        return request.path + "?" + urlencode(qs) + "#complaints"

    reset_url = request.path + "#complaints"
    prev_page_url = _page_url(complaints_page.prev_num) if complaints_page.has_prev else "#"
    next_page_url = _page_url(complaints_page.next_num) if complaints_page.has_next else "#"
 
    profile = None
    if session.get("role") == "student":
        profile = StudentProfile.query.filter_by(email=session["user"]).first()
        if not profile:
            display_name = session.get("full_name", "")
            profile = StudentProfile(
                email=session["user"],
                name=display_name,
                roll_number="",
                dept="",
                year="",
                phone=""
            )
            db.session.add(profile)
            db.session.commit()
 
        hostel = (profile.hostel_type or 'all').lower()
        blk = (profile.block or 'all').upper()
        notices = Notice.query.filter(
            Notice.hostel_type.in_(['all', 'ALL', hostel])
        ).filter(
            Notice.block.in_(['all', 'ALL', blk])
        ).order_by(Notice.created_at.desc()).all()
 
    return render_template(
        "welcome.html", email=session["user"], full_name=session.get("full_name", ""),
        role=session.get("role"), complaints=complaints, stats=stats,
        cat_labels=safe_json(cat_labels), cat_counts=safe_json(cat_counts),
        pri_dict=safe_json(pri_dict), month_labels=safe_json(month_labels),
        monthly_issued=safe_json(monthly_issued), monthly_resolved=safe_json(monthly_resolved),
        block_data=safe_json(block_data), block_status_data=safe_json(block_status_data),
        hostel_category_data=safe_json(hostel_category_data),
        hostel_totals=safe_json(hostel_totals),
        hostel_status_data=safe_json(hostel_status_data),
        hostel_block_data=safe_json(hostel_block_data),
        profile=profile,
        cat_labels_list=cat_labels,
        complaints_json=safe_json([c.to_dict() for c in complaints]),
        wardens=wardens,
        notices=notices,
        complaints_page=complaints_page,
        dashboard_scope=dashboard_scope,
        extended_analytics=extended_analytics,
        extended_analytics_json=safe_json(extended_analytics) if extended_analytics else "null",
        recent_activity=recent_activity,
        reset_url=reset_url,
        prev_page_url=prev_page_url,
        next_page_url=next_page_url,
        filters={
            "q": filter_args.get("q", ""),
            "status": filter_args.get("status", "all"),
            "category": filter_args.get("category", ""),
            "priority": filter_args.get("priority", ""),
            "hostel": filter_args.get("hostel", "all"),
            "block": filter_args.get("block", "all"),
            "date_from": filter_args.get("date_from", ""),
            "date_to": filter_args.get("date_to", ""),
            "sort": filter_args.get("sort", "latest"),
            "student_name": filter_args.get("student_name", ""),
            "room": filter_args.get("room", ""),
            "warden": filter_args.get("warden", ""),
        }
    )


@app.route("/welcome")
@login_required
def welcome():
    return welcome_view()


def _block_route_letter(letter):
    letter = (letter or "").upper()
    if letter not in VALID_BLOCKS:
        abort(404)
    return letter


@app.route("/admin/hostels")
@login_required
@admin_required
def admin_all_hostels():
    return welcome_view(forced_hostel="all", forced_block="all")


@app.route("/admin/boys")
@login_required
@admin_required
def admin_boys():
    return welcome_view(forced_hostel="boys", forced_block="all")


@app.route("/admin/boys/block-<letter>")
@login_required
@admin_required
def admin_boys_block(letter):
    return welcome_view(forced_hostel="boys", forced_block=_block_route_letter(letter))


@app.route("/admin/girls")
@login_required
@admin_required
def admin_girls():
    return welcome_view(forced_hostel="girls", forced_block="all")


@app.route("/admin/girls/block-<letter>")
@login_required
@admin_required
def admin_girls_block(letter):
    return welcome_view(forced_hostel="girls", forced_block=_block_route_letter(letter))


@app.route("/api/admin/dashboard")
@login_required
@admin_required
def api_admin_dashboard():
    """Reusable JSON API: GET /api/admin/dashboard?hostel=boys&block=A
    Returns everything the dashboard UI needs for a given hostel/block
    scope: summary statistics, category/priority/status distribution,
    weekly + monthly trends, heatmap data, block comparison, wardens
    summary, recent activity, and a paginated/searchable/sortable/filtered
    complaint list. Never mixes in data from other hostels/blocks.

    Complaint-list query params (all optional): page, per_page (<=50),
    q (search), status, category, priority, date_from, date_to, sort
    (same tokens the dashboard table headers use, e.g. 'priority_desc').
    """
    hostel = (request.args.get("hostel") or "all").lower()
    block = (request.args.get("block") or "all").upper()
    if hostel not in VALID_HOSTEL_TYPES:
        hostel = "all"
    if block not in VALID_BLOCKS:
        block = "all"

    stats = compute_scope_stats(hostel, block)
    if hostel == "all" and block == "all":
        stats["total"] = IssueCounter.get().total
    title, desc = dashboard_title_for(hostel, block)
    extended = compute_extended_analytics(hostel, block)

    q = _scope_query(hostel, block)

    from sqlalchemy import func as _func
    cat_rows = q.with_entities(Complaint.category, _func.count(Complaint.id)).group_by(Complaint.category).all()
    status_rows = q.with_entities(Complaint.status, _func.count(Complaint.id)).group_by(Complaint.status).all()
    pri_rows = q.with_entities(Complaint.priority, _func.count(Complaint.id)).group_by(Complaint.priority).all()

    # Wardens summary (scoped to hostel/block if given, else all wardens)
    wardens_scope = Warden.query.all()
    if hostel != "all":
        wardens_scope = [w for w in wardens_scope if w.hostel_type == hostel]
    if block != "all":
        wardens_scope = [w for w in wardens_scope if w.block == block]
    wardens_summary = [{"id": w.id, "name": w.name, "hostel_type": w.hostel_type, "block": w.block, "contact": w.contact} for w in wardens_scope]

    # Recent activity, scoped
    activity_q = ActivityLog.query
    if hostel != "all":
        activity_q = activity_q.filter(db.func.lower(ActivityLog.hostel_type) == hostel)
    if block != "all":
        activity_q = activity_q.filter(db.func.upper(ActivityLog.block) == block)
    recent_activity = activity_q.order_by(ActivityLog.created_at.desc()).limit(15).all()

    # Paginated / searchable / filterable / sortable complaint list.
    list_args = request.args.to_dict()
    list_args["hostel"] = hostel
    list_args["block"] = block
    filtered_q = apply_complaint_sort(
        apply_complaint_filters(Complaint.query, list_args),
        (request.args.get("sort") or "latest").strip(),
    )
    page = request.args.get("page", 1, type=int) or 1
    per_page = min(request.args.get("per_page", 10, type=int) or 10, 50)
    paginated = filtered_q.paginate(page=page, per_page=per_page, error_out=False)

    return {
        "scope": {"hostel": hostel, "block": block, "title": title, "description": desc},
        "stats": stats,
        "charts": {
            "by_category": {c: n for c, n in cat_rows},
            "by_status": {s: n for s, n in status_rows},
            "by_priority": {p: n for p, n in pri_rows},
            "weekly_trend": extended["weekly_trend"],
            "monthly_trend": extended["monthly_resolution"],
            "heatmap": extended["heatmap_data"],
            "block_comparison": extended["block_comparison"],
            "avg_resolution_trend": extended["avg_resolution_trend"],
            "frequent_locations": extended["frequent_locations"],
        },
        "wardens_summary": wardens_summary,
        "top_wardens": extended["top_wardens"],
        "recent_activities": [a.to_dict() for a in recent_activity],
        "recent_complaints": [c.to_dict() for c in paginated.items[:10]],
        "complaints": {
            "items": [c.to_dict() for c in paginated.items],
            "page": paginated.page,
            "pages": paginated.pages,
            "total": paginated.total,
            "per_page": per_page,
            "has_next": paginated.has_next,
            "has_prev": paginated.has_prev,
        },
    }


@app.route("/complaint")
@login_required
def complaint():
    if session.get("role") == "admin":
        flash("Admins cannot submit complaints.")
        return redirect(url_for("welcome"))
    profile = StudentProfile.query.filter_by(email=session["user"]).first()
    return render_template("complaint.html", full_name=session.get("full_name", ""), profile=profile)
 
 
@app.route("/submit_complaint", methods=["POST"])
@login_required
def submit_complaint():
    if session.get("role") == "admin":
        return redirect(url_for("welcome"))
 
    student_name = sanitize_string(request.form.get("name", ""), 120)
    room_number = sanitize_string(request.form.get("room", ""), 50)
    floor = sanitize_string(request.form.get("floor", ""), 10).strip()
    hostel_type_raw = (request.form.get("hostel_type") or "").strip().lower()
    block_raw = (request.form.get("block") or "").strip().upper()
    category_raw = (request.form.get("category") or "").strip()
    other_category_raw = (request.form.get("other_category") or "").strip()
    description = sanitize_description(request.form.get("description", ""), 500)
 
    hostel_type = hostel_type_raw if hostel_type_raw in VALID_HOSTEL_TYPES else None
    block = block_raw if block_raw in VALID_BLOCKS else None
    
    if category_raw == "Others":
        if other_category_raw:
            category = f"Others: {other_category_raw}"[:50]
        else:
            category = "Others"
    else:
        category = category_raw if category_raw in VALID_CATEGORIES else None
        
    priority = determine_priority(description)
 
    errors = []
    if not student_name:
        errors.append("Student name is required.")
    if not room_number:
        errors.append("Room number is required.")
    if not floor:
        errors.append("Floor details are required.")
    if not category:
        errors.append("Please select a valid category.")
    if not description:
        errors.append("Description is required.")
 
    if errors:
        for error in errors:
            flash(error, "error")
        return redirect(url_for("complaint"))
 
    c = Complaint(
        student_email=session["user"], student_name=student_name, room_number=room_number, floor=floor,
        hostel_type=hostel_type, block=block, category=category, priority=priority,
        description=description, status="Pending",
    )
    db.session.add(c)
    db.session.flush()
 
    counter = IssueCounter.get()
    counter.total += 1
    db.session.commit()
 
    uploaded = request.files.get("image")
    if uploaded and uploaded.filename and uploaded.filename.strip():
        if _allowed_image(uploaded.filename) and _is_valid_image_upload(uploaded):
            ext = uploaded.filename.rsplit(".", 1)[1].lower()
            final_filename = f"{c.id}_{secrets.token_hex(8)}.{ext}"
            uploaded.save(os.path.join(UPLOAD_DIR, final_filename))
            c.image_filename = final_filename
            db.session.commit()
        else:
            flash("Invalid image type. PNG, JPG, JPEG, WEBP allowed.", "warning")
 
    if genai_client:
        classify_priority_async(app, c.id, description)
 
    create_notification(
        ADMIN_EMAIL, "admin", "new_complaint",
        "New Complaint Submitted",
        f"{student_name} submitted a {category} complaint — Block {block or '?'}, Room {room_number}.",
        link="/welcome#complaints"
    )
    log_activity(
        "complaint_created",
        f"{student_name} raised a {category} complaint (#{c.id}) in {(hostel_type or '—').title()} Block {block or '—'}.",
        complaint=c, actor=session.get("user"),
    )
 
    flash("Complaint submitted successfully!", "success")
    return redirect(url_for("welcome", _anchor="complaints"))
 
 
@app.route("/update_profile", methods=["POST"])
@login_required
def update_profile():
    if session.get("role") != "student":
        abort(403)
        
    email = sanitize_string(request.form.get("email", ""), 120).lower().strip()
    name = sanitize_string(request.form.get("name", ""), 120).strip()
    roll_number = sanitize_string(request.form.get("roll_number", ""), 50).strip()
    dept = sanitize_string(request.form.get("dept", ""), 100).strip()
    year = sanitize_string(request.form.get("year", ""), 10).strip()
    phone = sanitize_string(request.form.get("phone", ""), 20).strip()
    hostel_type = sanitize_string(request.form.get("hostel_type", ""), 10).strip().lower()
    block = sanitize_string(request.form.get("block", ""), 10).strip().upper()
    floor = sanitize_string(request.form.get("floor", ""), 10).strip()
    room_no = sanitize_string(request.form.get("room_no", ""), 20).strip()
    
    if not (name and hostel_type and block and floor and room_no):
        flash("Name, hostel type, block, floor, and room number are required.", "error")
        return redirect(url_for("welcome"))

    if hostel_type not in VALID_HOSTEL_TYPES:
        flash("Please select a valid hostel type.", "error")
        return redirect(url_for("welcome"))

    if block not in VALID_BLOCKS:
        flash("Please select a valid block.", "error")
        return redirect(url_for("welcome"))
        
    if not email.endswith("@sece.ac.in"):
        flash("Official email must be a @sece.ac.in domain.", "error")
        return redirect(url_for("welcome"))
        
    profile = StudentProfile.query.filter_by(email=session["user"]).first()
    if not profile:
        profile = StudentProfile(email=session["user"])
        db.session.add(profile)
        
    if profile.email != email:
        existing = StudentProfile.query.filter_by(email=email).first()
        if existing:
            flash("Email is already registered to another profile.", "error")
            return redirect(url_for("welcome"))
            
        old_password = profile.password
        Complaint.query.filter_by(student_email=profile.email).update({Complaint.student_email: email})
        Notification.query.filter_by(recipient_email=profile.email).update({Notification.recipient_email: email})
        db.session.delete(profile)
        db.session.flush()
        
        profile = StudentProfile(
            email=email,
            name=name,
            roll_number=roll_number,
            dept=dept,
            year=year,
            phone=phone,
            password=old_password,
            hostel_type=hostel_type,
            block=block,
            floor=floor,
            room_no=room_no
        )
        db.session.add(profile)
        session["user"] = email
    else:
        profile.name = name
        profile.roll_number = roll_number
        profile.dept = dept
        profile.year = year
        profile.phone = phone
        profile.hostel_type = hostel_type
        profile.block = block
        profile.floor = floor
        profile.room_no = room_no
        
    session["full_name"] = name
    db.session.commit()
    flash("Profile updated successfully!", "success")
    return redirect(url_for("welcome"))


@app.route("/change_password", methods=["POST"])
@login_required
def change_password():
    if session.get("role") != "student":
        # The admin account is a single, env-var-managed credential rather
        # than a DB row, so there's nothing here to update in place.
        flash("The admin password is managed via the ADMIN_PASSWORD environment variable, not this form.", "error")
        return redirect(url_for("welcome"))

    current_password = request.form.get("current_password", "")
    new_password = request.form.get("new_password", "")
    confirm_password = request.form.get("confirm_password", "")

    profile = StudentProfile.query.filter_by(email=session["user"]).first()
    if not profile:
        flash("Profile not found.", "error")
        return redirect(url_for("welcome"))

    if not current_password or not new_password or not confirm_password:
        flash("All password fields are required.", "error")
        return redirect(url_for("welcome"))

    if not verify_and_maybe_upgrade_password(profile, current_password):
        flash("Current password is incorrect.", "error")
        return redirect(url_for("welcome"))

    if len(new_password) < 8:
        flash("New password must be at least 8 characters long.", "error")
        return redirect(url_for("welcome"))

    if new_password != confirm_password:
        flash("New password and confirmation do not match.", "error")
        return redirect(url_for("welcome"))

    if hmac.compare_digest(new_password, current_password):
        flash("New password must be different from your current password.", "error")
        return redirect(url_for("welcome"))

    profile.password = hash_password(new_password)
    db.session.commit()
    flash("Password changed successfully.", "success")
    return redirect(url_for("welcome"))


@app.route("/admin/reset_student_password", methods=["POST"])
@login_required
@admin_required
def admin_reset_student_password():
    """
    Stand-in for a full forgot-password flow: this app has no outbound
    email configured, so instead of leaving students unable to recover a
    lost password, an admin can issue a fresh temporary one from the
    student-details panel. The student is notified in-app and should
    change it immediately after logging back in.
    """
    email = sanitize_string(request.form.get("email", ""), 120).strip().lower()
    profile = StudentProfile.query.filter_by(email=email).first()
    if not profile:
        flash("Student not found.", "error")
        return redirect(url_for("welcome"))

    temp_password = secrets.token_urlsafe(9)
    profile.password = hash_password(temp_password)
    db.session.commit()

    create_notification(
        email, "student", "password_reset",
        "Password Reset by Admin",
        "Your password was reset by an administrator. Log in with the new "
        "temporary password shared with you and change it immediately from your profile.",
        link="/welcome"
    )

    flash(
        f"Password for {profile.name} ({email}) has been reset. "
        f"Temporary password: {temp_password} — share this with the student securely; "
        f"they should change it immediately after logging in.",
        "success"
    )
    return redirect(url_for("welcome"))
 
 
@app.route("/admin/student_details/<email>")
@login_required
@admin_required
def admin_student_details(email):
    profile = StudentProfile.query.filter_by(email=email).first()
    if not profile:
        return {"error": "Student profile not found"}, 404
    return {
        "name": profile.name,
        "email": profile.email,
        "roll_number": profile.roll_number or "—",
        "dept": profile.dept or "—",
        "year": profile.year or "—",
        "phone": profile.phone or "—",
        "hostel_type": (profile.hostel_type or "—").capitalize(),
        "block": profile.block or "—",
        "floor": profile.floor or "—",
        "room_no": profile.room_no or "—"
    }
 
 
@app.route("/admin/update_complaint/<int:cid>", methods=["POST"])
@login_required
@admin_required
def update_complaint(cid):
    c = Complaint.query.get_or_404(cid)
    new_status = (request.form.get("status") or "").strip()
    admin_note = sanitize_description(request.form.get("admin_note", ""), 500)
 
    status_changed = bool(new_status in VALID_STATUSES and new_status != c.status)
    if new_status in VALID_STATUSES:
        c.status = new_status
    c.admin_note = admin_note
    c.updated_at = datetime.utcnow()
    db.session.commit()

    if status_changed:
        create_notification(
            c.student_email, "student", "status_change",
            "Complaint Status Updated",
            f"Your complaint #{c.id} ({c.category}) is now '{c.status}'.",
            link="/welcome#complaints"
        )
        log_activity(
            "resolved" if c.status == "Resolved" else "status_updated",
            f"Complaint #{c.id} ({c.category}) status changed to '{c.status}' by admin.",
            complaint=c, actor=session.get("user"),
        )

    flash(f"Complaint #{cid} updated to '{c.status}'.", "success")
    return redirect(url_for("welcome"))
 
 
@app.route("/admin/delete_complaint/<int:cid>", methods=["POST"])
@login_required
@admin_required
def delete_complaint(cid):
    c = Complaint.query.get_or_404(cid)
    log_activity(
        "complaint_deleted",
        f"Complaint #{c.id} ({c.category}, {(c.hostel_type or '—').title()} Block {c.block or '—'}) was deleted by admin.",
        complaint=c, actor=session.get("user"),
    )
    if c.image_filename:
        img_path = os.path.join(UPLOAD_DIR, c.image_filename)
        if os.path.exists(img_path):
            try:
                os.remove(img_path)
            except OSError:
                pass
    db.session.delete(c)
    db.session.commit()
    flash(f"Complaint #{cid} deleted.", "success")
    return redirect(url_for("welcome"))


@app.route("/admin/bulk_update", methods=["POST"])
@login_required
@admin_required
def admin_bulk_update():
    """Applies one action (a status change or delete) to a batch of
    complaints selected via the table's row-selection checkboxes. Redirects
    back to wherever the request came from (whichever dashboard route)."""
    ids_raw = (request.form.get("ids") or "").strip()
    action = (request.form.get("action") or "").strip()
    redirect_to = request.form.get("redirect_to") or url_for("welcome")

    ids = []
    for part in ids_raw.split(","):
        part = part.strip()
        if part.isdigit():
            ids.append(int(part))

    if not ids:
        flash("No complaints were selected.", "warning")
        return redirect(redirect_to)

    complaints = Complaint.query.filter(Complaint.id.in_(ids)).all()
    if not complaints:
        flash("Selected complaints were not found.", "warning")
        return redirect(redirect_to)

    if action == "delete":
        count = 0
        for c in complaints:
            if c.image_filename:
                img_path = os.path.join(UPLOAD_DIR, c.image_filename)
                if os.path.exists(img_path):
                    try:
                        os.remove(img_path)
                    except OSError:
                        pass
            db.session.delete(c)
            count += 1
        db.session.commit()
        log_activity("complaint_deleted", f"Bulk-deleted {count} complaint(s) by admin.", actor=session.get("user"))
        flash(f"Deleted {count} complaint(s).", "success")
        return redirect(redirect_to)

    if action in VALID_STATUSES:
        count = 0
        for c in complaints:
            if c.status != action:
                c.status = action
                c.updated_at = datetime.utcnow()
                count += 1
                create_notification(
                    c.student_email, "student", "status_change",
                    "Complaint Status Updated",
                    f"Your complaint #{c.id} ({c.category}) is now '{c.status}'.",
                    link="/welcome#complaints"
                )
        db.session.commit()
        log_activity(
            "resolved" if action == "Resolved" else "status_updated",
            f"Bulk-updated {count} complaint(s) to '{action}' by admin.",
            actor=session.get("user"),
        )
        flash(f"Updated {count} complaint(s) to '{action}'.", "success")
        return redirect(redirect_to)

    flash("Invalid bulk action.", "error")
    return redirect(redirect_to)


@app.route("/admin/send_notification", methods=["POST"])
@login_required
@admin_required
def admin_send_notification():
    """Quick Actions -> Send Notification. Broadcasts a Notification to
    every student whose profile matches the given hostel/block scope
    ('all'/'all' broadcasts to every student with a profile)."""
    title = (request.form.get("title") or "").strip()
    message = (request.form.get("message") or "").strip()
    hostel = (request.form.get("hostel") or "all").strip().lower()
    block = (request.form.get("block") or "all").strip().upper()
    redirect_to = request.form.get("redirect_to") or url_for("welcome")

    if not title or not message:
        flash("Notification title and message are required.", "error")
        return redirect(redirect_to)

    recipients_q = StudentProfile.query
    if hostel != "all" and hostel in VALID_HOSTEL_TYPES:
        recipients_q = recipients_q.filter(db.func.lower(StudentProfile.hostel_type) == hostel)
    if block != "all" and block in VALID_BLOCKS:
        recipients_q = recipients_q.filter(db.func.upper(StudentProfile.block) == block)
    recipients = recipients_q.all()

    if not recipients:
        flash("No students matched that hostel/block scope — nothing sent.", "warning")
        return redirect(redirect_to)

    for student in recipients:
        create_notification(
            student.email, "student", "admin_broadcast", title, message[:500],
            link="/welcome#complaints"
        )

    scope_label = "All Students" if hostel == "all" and block == "all" else f"{hostel.title()}{' Block ' + block if block != 'all' else ''}"
    log_activity(
        "notice_sent",
        f"Notification \"{title}\" sent to {len(recipients)} student(s) in {scope_label}.",
        hostel_type=(None if hostel == "all" else hostel),
        block=(None if block == "all" else block),
        actor=session.get("user"),
    )
    flash(f"Notification sent to {len(recipients)} student(s).", "success")
    return redirect(redirect_to)


import csv
import io
 
def get_floor_sort_key(floor_str):
    if not floor_str:
        return (999999, "")
    val = floor_str.lower().strip()
    
    # Check common non-numeric terms first
    if "ground" in val or "g floor" in val or "g-floor" in val:
        return (0, val)
    if "basement" in val:
        return (-1, val)
        
    # Try to extract numbers
    nums = re.findall(r'\d+', val)
    if nums:
        return (int(nums[0]), val)
        
    # Text number checks
    if "first" in val or "1st" in val:
        return (1, val)
    if "second" in val or "2nd" in val:
        return (2, val)
    if "third" in val or "3rd" in val:
        return (3, val)
    if "fourth" in val or "4th" in val:
        return (4, val)
    if "fifth" in val or "5th" in val:
        return (5, val)
        
    # Fallback default sorting
    return (9999, val)
 
@app.route("/admin/reports")
@login_required
@admin_required
def admin_reports():
    return render_template(
        "reports.html",
        email=session["user"],
        full_name=session.get("full_name", ""),
        role=session.get("role")
    )
 
@app.route("/admin/reports/download")
@login_required
@admin_required
def admin_reports_download():
    hostel_type = request.args.get("hostel_type", "").strip().lower()
    report_type = request.args.get("report_type", "").strip().lower()
    block = request.args.get("block", "").strip().upper()
    category = request.args.get("category", "").strip()
    from_date_str = request.args.get("from_date", "").strip()
    to_date_str = request.args.get("to_date", "").strip()
 
    if hostel_type not in ["boys", "girls"]:
        return "Invalid hostel type specified", 400
 
    query = Complaint.query.filter(Complaint.hostel_type == hostel_type)
 
    if report_type == "each_block_all_categories":
        if not block:
            return "Block parameter is required for this report type", 400
        query = query.filter(Complaint.block == block)
 
    elif report_type == "each_block_each_category":
        if not block or not category:
            return "Both block and category parameters are required for this report type", 400
        query = query.filter(Complaint.block == block, Complaint.category == category)
 
    elif report_type == "all_blocks_each_category":
        if not category:
            return "Category parameter is required for this report type", 400
        query = query.filter(Complaint.category == category)
 
    elif report_type == "all_blocks_all_categories":
        # No extra filters
        pass
    else:
        return "Invalid report type specified", 400
 
    if from_date_str:
        try:
            from_dt = datetime.strptime(from_date_str, "%Y-%m-%d")
            query = query.filter(Complaint.created_at >= from_dt)
        except ValueError:
            pass
 
    if to_date_str:
        try:
            to_dt = datetime.strptime(to_date_str, "%Y-%m-%d")
            to_dt = to_dt.replace(hour=23, minute=59, second=59, microsecond=999999)
            query = query.filter(Complaint.created_at <= to_dt)
        except ValueError:
            pass
 
    complaints = query.all()
    if not complaints:
        flash("No complaints registered on selected period of time", "warning")
        return redirect(url_for("admin_reports"))
 
    # Sort logic: block A to F, then each block details must be floor 1 to last
    complaints_sorted = sorted(
        complaints,
        key=lambda c: (c.block or "", get_floor_sort_key(c.floor))
    )
 
    from flask import Response
    
    def generate_csv():
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write headers
        writer.writerow([
            "Complaint ID",
            "Student Name",
            "Student Email",
            "Hostel Type",
            "Block",
            "Floor",
            "Room Number",
            "Category",
            "Priority",
            "Description",
            "Status",
            "Admin Note",
            "Created At",
            "Updated At"
        ])
        yield output.getvalue()
        output.truncate(0)
        output.seek(0)
 
        for c in complaints_sorted:
            writer.writerow([
                c.id,
                _csv_safe(c.student_name),
                _csv_safe(c.student_email),
                (c.hostel_type or "").capitalize(),
                c.block or "—",
                c.floor or "—",
                _csv_safe(c.room_number),
                _csv_safe(c.category),
                c.priority,
                _csv_safe(c.description),
                c.status,
                _csv_safe(c.admin_note) if c.admin_note else "—",
                c.created_at.strftime('%Y-%m-%d %H:%M:%S') if c.created_at else "—",
                c.updated_at.strftime('%Y-%m-%d %H:%M:%S') if c.updated_at else "—"
            ])
            yield output.getvalue()
            output.truncate(0)
            output.seek(0)
 
    filename_parts = [hostel_type, report_type]
    if block:
        filename_parts.append(f"block_{block}")
    if category:
        filename_parts.append(category.replace(" ", "_").lower())
    
    filename = f"hostel_report_{'_'.join(filename_parts)}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.csv"
 
    response = Response(generate_csv(), mimetype="text/csv")
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return response
 
 
@app.route("/admin/reports/view")
@login_required
@admin_required
def admin_reports_view():
    hostel_type = request.args.get("hostel_type", "").strip().lower()
    report_type = request.args.get("report_type", "").strip().lower()
    block = request.args.get("block", "").strip().upper()
    category = request.args.get("category", "").strip()
    from_date_str = request.args.get("from_date", "").strip()
    to_date_str = request.args.get("to_date", "").strip()
 
    if hostel_type not in ["boys", "girls"]:
        return "Invalid hostel type specified", 400
 
    query = Complaint.query.filter(Complaint.hostel_type == hostel_type)
 
    if report_type == "each_block_all_categories":
        if not block:
            return "Block parameter is required for this report type", 400
        query = query.filter(Complaint.block == block)
 
    elif report_type == "each_block_each_category":
        if not block or not category:
            return "Both block and category parameters are required for this report type", 400
        query = query.filter(Complaint.block == block, Complaint.category == category)
 
    elif report_type == "all_blocks_each_category":
        if not category:
            return "Category parameter is required for this report type", 400
        query = query.filter(Complaint.category == category)
 
    elif report_type == "all_blocks_all_categories":
        pass
    else:
        return "Invalid report type specified", 400
 
    if from_date_str:
        try:
            from_dt = datetime.strptime(from_date_str, "%Y-%m-%d")
            query = query.filter(Complaint.created_at >= from_dt)
        except ValueError:
            pass
 
    if to_date_str:
        try:
            to_dt = datetime.strptime(to_date_str, "%Y-%m-%d")
            to_dt = to_dt.replace(hour=23, minute=59, second=59, microsecond=999999)
            query = query.filter(Complaint.created_at <= to_dt)
        except ValueError:
            pass
 
    complaints = query.all()
    if not complaints:
        flash("No complaints registered on selected period of time", "warning")
        return redirect(url_for("admin_reports"))
 
    # Sort logic: block A to F, then each block details must be floor 1 to last
    complaints_sorted = sorted(
        complaints,
        key=lambda c: (c.block or "", get_floor_sort_key(c.floor))
    )
 
    scope_str = "All Blocks with All Categories"
    if report_type == "each_block_all_categories":
        scope_str = f"Block {block} with All Categories"
    elif report_type == "each_block_each_category":
        scope_str = f"Block {block} with {category}"
    elif report_type == "all_blocks_each_category":
        scope_str = f"All Blocks with {category}"
 
    return render_template(
        "report_print.html",
        complaints=complaints_sorted,
        hostel_type=hostel_type,
        scope_str=scope_str,
        generation_date=datetime.utcnow().strftime('%B %d, %Y at %I:%M %p UTC'),
        admin_name=session.get("full_name", session["user"])
    )


# ─────────────────────────────────────────────────────────────────────────
# EXPORT & SCOPED REPORT GENERATION (CSV / Excel / PDF)
# ─────────────────────────────────────────────────────────────────────────
# Two related-but-distinct features:
#   /admin/export/<fmt>          — exports exactly the currently filtered/
#                                   searched complaint list shown on a
#                                   dashboard page (all matching rows,
#                                   independent of pagination).
#   /admin/dashboard-report/<fmt> — a full statistics report (summary,
#                                   breakdowns, monthly trend, table) scoped
#                                   to whichever hostel/block dashboard page
#                                   it was generated from.
# Both are additive — they don't touch the existing /admin/reports flow.

EXPORT_COLUMNS = [
    "Complaint ID", "Student Name", "Roll Number", "Hostel", "Block",
    "Room Number", "Category", "Priority", "Status", "Created Date", "Assigned Warden",
]


def _export_rows_for(complaints, warden_by_scope, profiles_by_email):
    rows = []
    for c in complaints:
        prof = profiles_by_email.get(c.student_email)
        w = warden_by_scope.get(((c.hostel_type or "").lower(), (c.block or "").upper()))
        rows.append([
            c.id,
            c.student_name,
            (prof.roll_number if prof and prof.roll_number else "—"),
            (c.hostel_type or "—").capitalize(),
            c.block or "—",
            c.room_number,
            c.category,
            c.priority,
            c.status,
            c.created_at.strftime('%Y-%m-%d %H:%M') if c.created_at else "—",
            w.name if w else "Unassigned",
        ])
    return rows


def _fetch_complaints_and_lookups(query):
    complaints = query.order_by(Complaint.created_at.desc()).all()
    emails = {c.student_email for c in complaints}
    profiles_by_email = {p.email: p for p in StudentProfile.query.filter(StudentProfile.email.in_(emails)).all()} if emails else {}
    warden_by_scope = {(w.hostel_type, w.block): w for w in Warden.query.all()}
    return complaints, profiles_by_email, warden_by_scope


def _csv_download(rows, columns, filename):
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(columns)
    for row in rows:
        writer.writerow([_csv_safe(v) for v in row])
    resp = Response(output.getvalue(), mimetype="text/csv")
    resp.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return resp


def _xlsx_download(rows, columns, filename, sheet_title="Complaints", extra_sheets=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]

    header_fill = PatternFill(start_color="0A2472", end_color="0A2472", fill_type="solid")
    header_font = Font(color="FFC107", bold=True)
    ws.append(columns)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append(row)
    for i, col in enumerate(columns, start=1):
        max_len = max([len(str(col))] + [len(str(r[i - 1])) for r in rows]) if rows else len(str(col))
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 4, 40)
    ws.freeze_panes = "A2"

    if extra_sheets:
        for title, header, data_rows in extra_sheets:
            s = wb.create_sheet(title[:31])
            s.append(header)
            for cell in s[1]:
                cell.font = Font(bold=True)
            for r in data_rows:
                s.append(r)
            for i, col in enumerate(header, start=1):
                s.column_dimensions[get_column_letter(i)].width = max(len(str(col)) + 4, 14)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def _pdf_download(filename, title, subtitle, sections, table_columns, table_rows):
    """sections: list of (heading, [(label, value), ...]) rendered as small
    key/value blocks before the main complaint table."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        topMargin=16 * mm, bottomMargin=14 * mm, leftMargin=14 * mm, rightMargin=14 * mm,
    )
    styles = getSampleStyleSheet()
    navy = colors.HexColor("#0A2472")
    gold = colors.HexColor("#FFC107")
    title_style = ParagraphStyle("TitleStyle", parent=styles["Title"], textColor=navy, fontSize=18)
    sub_style = ParagraphStyle("SubStyle", parent=styles["Normal"], textColor=colors.HexColor("#555555"), fontSize=10)
    heading_style = ParagraphStyle("HeadingStyle", parent=styles["Heading3"], textColor=navy, spaceBefore=10, spaceAfter=4)

    elements = [Paragraph(title, title_style), Paragraph(subtitle, sub_style), Spacer(1, 10)]

    for heading, kv_pairs in sections:
        elements.append(Paragraph(heading, heading_style))
        data = [[label, str(value)] for label, value in kv_pairs]
        # Lay key/value pairs out three per row for compactness
        grid_rows = [data[i:i + 3] for i in range(0, len(data), 3)]
        flat = [[cell for pair in row for cell in pair] for row in grid_rows]
        t = Table(flat, hAlign="LEFT")
        t.setStyle(TableStyle([
            ("FONTSIZE", (0, 0), (-1, -1), 8.5),
            ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#333333")),
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        elements.append(t)
        elements.append(Spacer(1, 6))

    elements.append(Paragraph("Complaint Table", heading_style))
    table_data = [table_columns] + [[str(v) for v in row] for row in table_rows]
    ctable = Table(table_data, repeatRows=1, hAlign="LEFT")
    ctable.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), navy),
        ("TEXTCOLOR", (0, 0), (-1, 0), gold),
        ("FONTSIZE", (0, 0), (-1, -1), 7.5),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#DDDDDD")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F7FB")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elements.append(ctable)

    doc.build(elements)
    buf.seek(0)
    return Response(
        buf.read(),
        mimetype="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.route("/admin/export/<fmt>")
@login_required
@admin_required
def admin_export(fmt):
    """Exports exactly the currently filtered/searched complaints (all
    matching rows, ignoring pagination) shown on whichever dashboard page
    this was triggered from."""
    if fmt not in ("csv", "xlsx", "pdf"):
        abort(404)

    query = apply_complaint_sort(
        apply_complaint_filters(Complaint.query, request.args),
        (request.args.get("sort") or "latest").strip()
    )
    complaints, profiles_by_email, warden_by_scope = _fetch_complaints_and_lookups(query)
    rows = _export_rows_for(complaints, warden_by_scope, profiles_by_email)

    hostel = (request.args.get("hostel") or "all").lower()
    block = (request.args.get("block") or "all").upper()
    scope_slug = hostel if hostel != "all" else "all-hostels"
    if block != "all":
        scope_slug += f"_block-{block}"
    stamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    base_name = f"spark_desk_complaints_{scope_slug}_{stamp}"

    if not rows:
        flash("No complaints match the current filters — nothing to export.", "warning")
        return redirect(request.referrer or url_for("welcome"))

    if fmt == "csv":
        return _csv_download(rows, EXPORT_COLUMNS, f"{base_name}.csv")
    if fmt == "xlsx":
        return _xlsx_download(rows, EXPORT_COLUMNS, f"{base_name}.xlsx")
    title, desc = dashboard_title_for(hostel, block)
    return _pdf_download(
        f"{base_name}.pdf",
        f"SPARK Desk — Complaint Export",
        f"{title} · Generated {datetime.utcnow().strftime('%d %b %Y, %I:%M %p UTC')} · {len(rows)} complaint(s)",
        sections=[],
        table_columns=EXPORT_COLUMNS,
        table_rows=rows,
    )


@app.route("/admin/dashboard-report/<fmt>")
@login_required
@admin_required
def admin_dashboard_report(fmt):
    """Generates a full statistics report — summary, resolution rate,
    category/priority/status breakdowns, 6-month trend, and the complaint
    table — scoped to the hostel/block of the dashboard page it was
    triggered from. Never mixes in data from other hostels/blocks."""
    if fmt not in ("csv", "xlsx", "pdf"):
        abort(404)

    hostel = (request.args.get("hostel") or "all").lower()
    block = (request.args.get("block") or "all").upper()
    if hostel not in VALID_HOSTEL_TYPES:
        hostel = "all"
    if block not in VALID_BLOCKS:
        block = "all"

    q = Complaint.query
    if hostel != "all":
        q = q.filter(db.func.lower(Complaint.hostel_type) == hostel)
    if block != "all":
        q = q.filter(db.func.upper(Complaint.block) == block)

    stats = compute_scope_stats(hostel, block)
    if hostel == "all" and block == "all":
        stats["total"] = IssueCounter.get().total
    title, desc = dashboard_title_for(hostel, block)

    from sqlalchemy import func as _func
    cat_rows = q.with_entities(Complaint.category, _func.count(Complaint.id)).group_by(Complaint.category).order_by(_func.count(Complaint.id).desc()).all()
    status_rows = dict(q.with_entities(Complaint.status, _func.count(Complaint.id)).group_by(Complaint.status).all())
    pri_rows = dict(q.with_entities(Complaint.priority, _func.count(Complaint.id)).group_by(Complaint.priority).all())

    now = datetime.utcnow()
    monthly_trend = []
    for i in range(5, -1, -1):
        m = now.month - i
        y = now.year
        while m <= 0:
            m += 12
            y -= 1
        month_start = datetime(y, m, 1)
        month_end = datetime(y + 1, 1, 1) if m == 12 else datetime(y, m + 1, 1)
        count = q.filter(Complaint.created_at >= month_start, Complaint.created_at < month_end).count()
        monthly_trend.append((datetime(y, m, 1).strftime("%b %Y"), count))

    resolution_rate = round((stats["resolved"] / stats["active"]) * 100, 1) if stats["active"] else 0

    complaints, profiles_by_email, warden_by_scope = _fetch_complaints_and_lookups(q)
    rows = _export_rows_for(complaints, warden_by_scope, profiles_by_email)

    stamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    scope_slug = hostel if hostel != "all" else "all-hostels"
    if block != "all":
        scope_slug += f"_block-{block}"
    base_name = f"spark_desk_report_{scope_slug}_{stamp}"

    summary_kv = [
        ("Total Complaints", stats["total"]),
        ("Pending", stats["pending"]),
        ("In Progress", stats["in_progress"]),
        ("Resolved", stats["resolved"]),
        ("Resolution Rate", f"{resolution_rate}%"),
        ("High Priority", stats["high"]),
        ("Medium Priority", stats["medium"]),
        ("Low Priority", stats["low"]),
        ("Today's Complaints", stats["today"]),
        ("Resolved Today", stats["resolved_today"]),
        ("Avg Resolution Time", f"{stats['avg_resolution_hours']}h"),
        ("Monthly Growth", f"{stats['monthly_growth']}%"),
    ]
    category_kv = [(cat, n) for cat, n in cat_rows] or [("No data", 0)]
    priority_kv = [("High", pri_rows.get("High", 0)), ("Moderate", pri_rows.get("Moderate", 0)), ("Low", pri_rows.get("Low", 0))]
    status_kv = [("Pending", status_rows.get("Pending", 0)), ("In Progress", status_rows.get("In Progress", 0)), ("Resolved", status_rows.get("Resolved", 0))]
    trend_kv = monthly_trend

    if fmt == "csv":
        # CSV can't hold sections cleanly — prepend a compact summary block
        # above the complaint table in the same file.
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow([f"SPARK Desk Report — {title}"])
        writer.writerow([f"Generated {datetime.utcnow().strftime('%d %b %Y, %I:%M %p UTC')}"])
        writer.writerow([])
        writer.writerow(["Summary"])
        for label, value in summary_kv:
            writer.writerow([label, value])
        writer.writerow([])
        writer.writerow(["Category Breakdown"])
        for label, value in category_kv:
            writer.writerow([label, value])
        writer.writerow([])
        writer.writerow(["Priority Breakdown"])
        for label, value in priority_kv:
            writer.writerow([label, value])
        writer.writerow([])
        writer.writerow(["Status Breakdown"])
        for label, value in status_kv:
            writer.writerow([label, value])
        writer.writerow([])
        writer.writerow(["Monthly Trend (last 6 months)"])
        for label, value in trend_kv:
            writer.writerow([label, value])
        writer.writerow([])
        writer.writerow(["Complaint Table"])
        writer.writerow(EXPORT_COLUMNS)
        for row in rows:
            writer.writerow([_csv_safe(v) for v in row])
        resp = Response(output.getvalue(), mimetype="text/csv")
        resp.headers["Content-Disposition"] = f"attachment; filename={base_name}.csv"
        return resp

    if fmt == "xlsx":
        extra_sheets = [
            ("Summary", ["Metric", "Value"], summary_kv),
            ("Categories", ["Category", "Count"], category_kv),
            ("Priority", ["Priority", "Count"], priority_kv),
            ("Status", ["Status", "Count"], status_kv),
            ("Monthly Trend", ["Month", "Count"], trend_kv),
        ]
        return _xlsx_download(rows, EXPORT_COLUMNS, f"{base_name}.xlsx", sheet_title="Complaints", extra_sheets=extra_sheets)

    return _pdf_download(
        f"{base_name}.pdf",
        f"SPARK Desk — Report: {title}",
        f"{desc} · Generated {datetime.utcnow().strftime('%d %b %Y, %I:%M %p UTC')}",
        sections=[
            ("Summary", summary_kv),
            ("Category Breakdown", category_kv),
            ("Priority Breakdown", priority_kv),
            ("Status Breakdown", status_kv),
            ("Monthly Trend (last 6 months)", trend_kv),
        ],
        table_columns=EXPORT_COLUMNS,
        table_rows=rows,
    )


@app.route("/admin/wardens")
@login_required
@admin_required
def admin_wardens():
    wardens = Warden.query.all()
    return render_template(
        "wardens.html",
        email=session["user"],
        full_name=session.get("full_name", ""),
        role=session.get("role"),
        wardens=wardens
    )
 
 
@app.route("/admin/wardens/edit", methods=["POST"])
@login_required
@admin_required
def admin_wardens_edit():
    warden_id = request.form.get("warden_id")
    name = sanitize_string(request.form.get("name", ""), 120).strip()
    contact = sanitize_string(request.form.get("contact", ""), 50).strip()
 
    if not warden_id or not name or not contact:
        flash("All fields are required.", "error")
        return redirect(url_for("welcome"))

    if not warden_id.isdigit():
        flash("Invalid warden reference.", "error")
        return redirect(url_for("admin_wardens"))

    warden = Warden.query.get_or_404(int(warden_id))
    warden.name = name
    warden.contact = contact
    try:
        db.session.commit()
        flash(f"Warden details for Block {warden.block} ({warden.hostel_type.capitalize()}) updated.", "success")
        log_activity(
            "warden_updated",
            f"Warden for {warden.hostel_type.title()} Block {warden.block} updated to {warden.name}.",
            hostel_type=warden.hostel_type, block=warden.block, actor=session.get("user"),
        )
    except Exception as e:
        db.session.rollback()
        print(f"[HOSTEL APP] Failed to update warden {warden_id}: {e}")
        flash("Failed to update warden. Please try again.", "error")
 
    return redirect(url_for("admin_wardens"))
 
 
@app.route("/admin/notices")
@login_required
@admin_required
def admin_notices():
    notices = Notice.query.order_by(Notice.created_at.desc()).all()
    return render_template(
        "notices.html",
        email=session["user"],
        full_name=session.get("full_name", ""),
        role=session.get("role"),
        notices=notices
    )
 
 
@app.route("/admin/notices/add", methods=["POST"])
@login_required
@admin_required
def admin_notices_add():
    hostel_type = sanitize_string(request.form.get("hostel_type", ""), 10).strip().lower()
    if hostel_type not in ("boys", "girls", "all"):
        flash("Invalid hostel type selected.", "error")
        return redirect(url_for("admin_notices"))
    block = sanitize_string(request.form.get("block", ""), 10).strip()
    if block.lower() == "all":
        block = "all"
    else:
        block = block.upper()
    title = sanitize_string(request.form.get("title", ""), 120).strip()
    content = sanitize_string(request.form.get("content", ""), 1000).strip()
    resolved_by_str = request.form.get("resolved_by", "").strip()
 
    if not hostel_type or not block or not title or not content:
        flash("All fields are required.", "error")
        return redirect(url_for("admin_notices"))
 
    resolved_by = None
    if resolved_by_str:
        try:
            resolved_by = datetime.strptime(resolved_by_str, "%Y-%m-%dT%H:%M")
        except ValueError:
            pass
 
    try:
        new_notice = Notice(
            hostel_type=hostel_type,
            block=block,
            title=title,
            content=content,
            resolved_by=resolved_by
        )
        db.session.add(new_notice)
        db.session.commit()
        log_activity(
            "notice_sent",
            f"Notice \"{title}\" sent to {hostel_type.title() if hostel_type != 'all' else 'All Hostels'}"
            f"{' Block ' + block if block != 'all' else ''}.",
            hostel_type=(None if hostel_type == "all" else hostel_type),
            block=(None if block == "all" else block),
            actor=session.get("user"),
        )

        # Notify every student the notice applies to (same matching rule
        # used to display notices on their dashboard: hostel_type/block
        # 'all' matches everyone).
        recipients_query = StudentProfile.query
        if hostel_type != "all":
            recipients_query = recipients_query.filter(
                db.or_(StudentProfile.hostel_type == hostel_type, StudentProfile.hostel_type == None)
            )
        if block != "all":
            recipients_query = recipients_query.filter(
                db.or_(StudentProfile.block == block, StudentProfile.block == None)
            )
        for student in recipients_query.all():
            create_notification(
                student.email, "student", "new_notice",
                "New Notice",
                title,
                link="/welcome#notices"
            )

        flash("Notice added successfully.", "success")
    except Exception as e:
        db.session.rollback()
        print(f"[HOSTEL APP] Failed to add notice: {e}")
        flash("Failed to add notice. Please try again.", "error")
 
    return redirect(url_for("admin_notices"))
 
 
@app.route("/admin/notices/delete/<int:nid>", methods=["POST"])
@login_required
@admin_required
def admin_notices_delete(nid):
    notice = Notice.query.get_or_404(nid)
    try:
        db.session.delete(notice)
        db.session.commit()
        flash("Notice deleted successfully.", "success")
    except Exception as e:
        db.session.rollback()
        print(f"[HOSTEL APP] Failed to delete notice {nid}: {e}")
        flash("Failed to delete notice. Please try again.", "error")
 
    return redirect(url_for("admin_notices"))


# --- Notifications --------------------------------------------------------
NOTIFICATIONS_PAGE_SIZE = 20


@app.route("/notifications")
@login_required
def notifications_page():
    page = request.args.get("page", 1, type=int)
    if page < 1:
        page = 1
    pagination = Notification.query.filter_by(
        recipient_email=session["user"]
    ).order_by(Notification.created_at.desc()).paginate(
        page=page, per_page=NOTIFICATIONS_PAGE_SIZE, error_out=False
    )
    return render_template(
        "notifications.html",
        email=session["user"], full_name=session.get("full_name", ""),
        role=session.get("role"), pagination=pagination,
        notifications=pagination.items
    )


@app.route("/api/notifications/unread_count")
@login_required
def api_notifications_unread_count():
    count = Notification.query.filter_by(recipient_email=session["user"], is_read=False).count()
    return {"count": count}


@app.route("/api/notifications/recent")
@login_required
def api_notifications_recent():
    items = Notification.query.filter_by(
        recipient_email=session["user"]
    ).order_by(Notification.created_at.desc()).limit(8).all()
    return {"notifications": [n.to_dict() for n in items]}


@app.route("/notifications/read/<int:nid>", methods=["POST"])
@login_required
def notifications_mark_read(nid):
    note = Notification.query.filter_by(id=nid, recipient_email=session["user"]).first_or_404()
    note.is_read = True
    db.session.commit()
    if request.headers.get("X-Requested-With") == "fetch":
        return {"ok": True}
    return redirect(note.link or url_for("notifications_page"))


@app.route("/notifications/read_all", methods=["POST"])
@login_required
def notifications_mark_all_read():
    Notification.query.filter_by(recipient_email=session["user"], is_read=False).update({"is_read": True})
    db.session.commit()
    if request.headers.get("X-Requested-With") == "fetch":
        return {"ok": True}
    flash("All notifications marked as read.", "success")
    return redirect(url_for("notifications_page"))


@app.route("/logout")
def logout():
 
    session.clear()
    flash("You have been logged out.", "info")
    return redirect(url_for("login"))
 
 
@app.errorhandler(400)
def bad_request(e):
    message = getattr(e, "description", None) or "The request could not be understood or was invalid."
    return render_template(
        "error.html", code=400, title="Bad Request",
        message=message,
        show_login_link="user" not in session
    ), 400


@app.errorhandler(401)
def unauthorized(e):
    return render_template(
        "error.html", code=401, title="Sign In Required",
        message="You need to be logged in to view this page.",
        show_login_link=True
    ), 401


@app.errorhandler(403)
def forbidden(e):
    return render_template(
        "error.html", code=403, title="Access Denied",
        message="You don't have permission to view this page.",
        show_login_link="user" not in session
    ), 403


@app.errorhandler(404)
def not_found(e):
    return render_template(
        "error.html", code=404, title="Page Not Found",
        message="The page you're looking for doesn't exist or may have moved.",
        show_login_link="user" not in session
    ), 404


@app.errorhandler(405)
def method_not_allowed(e):
    return render_template(
        "error.html", code=405, title="Method Not Allowed",
        message="That action isn't supported for this page.",
        show_login_link="user" not in session
    ), 405


@app.errorhandler(500)
def server_error(e):
    db.session.rollback()
    return render_template(
        "error.html", code=500, title="Something Went Wrong",
        message="An unexpected server error occurred. Please try again in a moment.",
        show_login_link="user" not in session
    ), 500
 
 
if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=int(os.environ.get("PORT", 5000)))
